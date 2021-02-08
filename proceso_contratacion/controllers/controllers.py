# -*- coding: utf-8 -*-
from odoo import http
from datetime import datetime
from docxtpl import DocxTemplate
from jinja2 import Template
import json
import time
from num2words import num2words
import binascii #nuevo


from datetime import datetime, date, time, timedelta
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from openpyxl.styles import colors
from openpyxl.styles import Font, Color, NamedStyle
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

class ProcesoContratacion(http.Controller):
    @http.route('/documento/CONTRATO/<contrato>/idplantilla/<id>', type='http', auth='public', website=True)
    def contrato(self, contrato, id):
        try:


            busca_platilla = http.request.env['plantillas.plantillas'].sudo().browse(int(id))[0] #nuevo

            xd = binascii.a2b_base64(busca_platilla.subir_documento) #nuevo
            with open("/tmp/doc2.doc", 'wb') as file: #nuevo

                file.write(xd) #nuevo
            doc = DocxTemplate("/tmp/doc2.doc") #nuevo
            

            qwerty = http.request.env['proceso.elaboracion_contrato'].sudo().browse(int(contrato))[0]
            for i in qwerty:

                adjudicacion = i.adjudicacion
                contrato = i.contrato  # Número de contrato
                fecha = i.fecha.strftime("%d/%m/%Y")  # Fecha actual
                descripcion_meta = i.name
                descripcion_trabajos = i.descripciontrabajos
                supervicion_externa = i.supervisionexterna
                fecha_inicio = i.fechainicio.strftime("%d/%m/%Y")
                fecha_termino = i.fechatermino.strftime("%d/%m/%Y")
                adjudicacion = i.adjudicacion.numerocontrato
                contratista = i.contratista.name
                representante_contratista = i.contratista.nombre_representante
                calle = i.contratista.calles
                colonia = i.contratista.colonia
                numero = i.contratista.numero
                estado = i.contratista.estado_entidad.estado
                municipio = i.contratista.municipio_delegacion
                codigo_postal = i.contratista.cp

                periodo_retencion = i.periodicidadretencion
                unidad_resposable = i.unidadresponsableejecucion.name
                retencion = i.retencion
                tipo_contrato = i.tipo_contrato

                if tipo_contrato == "1":
                    for hr in i.contrato_partida_licitacion:
                        oficio = hr.recursos.name.name
                        fecha_oficio_ = hr.recursos.name.fecha_actual

                    tipo_contrato = "Licitación"
                elif tipo_contrato == "2":
                    for rf in i.contrato_partida_adjudicacion:
                        oficio = rf.recursos.name.name  # Número de oficio
                        fecha_oficio_ = rf.recursos.name.fecha_actual  # Fecha de cuando se realizo el oficio
                    tipo_contrato = "Adjudicación"

            obra_partida_descripcion = ""
            for o in qwerty.contrato_partida_adjudicacion:
                obra_partida_descripcion = obra_partida_descripcion + " " + o.obra.descripcion
                obra = o.obra.descripcion
                monto_partida = o.monto_partida
                monto_partida_letra = num2words(monto_partida, lang='es')

                iva_partida = o.iva_partida
                iva_partida_letra = num2words(iva_partida, lang='es')

                total_partida = o.total_partida
                total_partida_letra = num2words(total_partida, lang='es')

            partidas = http.request.env['partidas.partidas'].sudo().browse(int(o))[0]
            for pp in partidas.programaInversion:
                programa_inversion = pp.name

            ejercicio = http.request.env['registro.ejercicio'].sudo().browse(int(i.ejercicio))[0]
            for ej in ejercicio:
                ejercicio = ej.name

            meses = ["Unknown",
                     "Enero",
                     "Febrero",
                     "Marzo",
                     "Abril",
                     "Mayo",
                     "Junio",
                     "Julio",
                     "Agosto",
                     "Septiembre",
                     "Octubre",
                     "Noviembre",
                     "Diciembre"] 

            fecha_oficio_letra = str(fecha_oficio_.day) + " de " + meses[fecha_oficio_.month] + " de " + str(fecha_oficio_.year)
            #Firmas
            titular = http.request.env['ir.config_parameter'].sudo().get_param('firmas_logos.nombre_titular')
            testigo_1 = http.request.env['ir.config_parameter'].sudo().get_param('firmas_logos.nombre_testigo_uno')
            testigo_2 = http.request.env['ir.config_parameter'].sudo().get_param('firmas_logos.nombre_testigo_dos')
            testigo_3 = http.request.env['ir.config_parameter'].sudo().get_param('firmas_logos.nombre_testigo_tres')
            puesto_testigo_1 = http.request.env['ir.config_parameter'].sudo().get_param('firmas_logos.puesto_uno')
            puesto_testigo_2 = http.request.env['ir.config_parameter'].sudo().get_param('firmas_logos.puesto_dos')
            puesto_testigo_3 = http.request.env['ir.config_parameter'].sudo().get_param('firmas_logos.puesto_tres')

            context = {
                'ejercicio': ejercicio,  #Año de la Fecha Inicial del Contrato
                'contrato': contrato,  #Número de Contrato
                'tipo_contrato': tipo_contrato,  #Objeto del Contrato
                'monto_partida': '{:20,.2f}'.format(monto_partida), #Monto del Contrato Sin IVA
                'iva_partida': '{:20,.2f}'.format(iva_partida),  #IVA de la partida
                'total_partida': '{:20,.2f}'.format(total_partida),  #Importe del contrato con IVA
                'contratista': contratista,  #Nombre de la Empresa Contratista
                'obra': obra,
                'calle': calle,  #Calle del Domicilio de la Empresa Contratista
                'colonia': colonia,  #Colonia del Domicilio de la Empresa Contratista
                'numero': numero,  #Número del Domicilio de la Empresa Contratista
                'estado': estado,  #Estado del Domicilio de la Empresa Contratista
                'municipio': municipio,  #Municipio del Domicilio de la Empresa Contratista
                'codigo_postal': codigo_postal,  #Código Postal del Domicilio de la Empresa Contratista
                'fecha': fecha,  #Fecha del contrato
                'fecha_inicio': fecha_inicio,  #Fecha de Inicio del Contrato
                'fecha_termino': fecha_termino,  #Fecha Final del Contrato
                'adjudicacion': adjudicacion,  # Adjudicación que dio origen al contrato



                'descripcion_meta': descripcion_meta,
                'descripcion_trabajos': descripcion_trabajos,
                'supervicion_externa': supervicion_externa,
                'programa_inversion': programa_inversion,
                'partidas': obra_partida_descripcion,
                'periodo_retencion': periodo_retencion,
                'unidad_resposable': unidad_resposable,
                'retencion': retencion,
                'monto_partida_letra': monto_partida_letra,
                'total_partida_letra': total_partida_letra,
                'iva_partida_letra': iva_partida_letra,

                'numero_oficio': oficio,
                'fecha_oficio': fecha_oficio_,
                'fecha_oficio_letra': fecha_oficio_letra,

                #Firmas
                'titular': titular,
                'nombre_representante_contratista': representante_contratista,
                'testigo_1': testigo_1,
                'testigo_2': testigo_2,
                'testigo_3': testigo_3,
                'puesto_testigo_1': puesto_testigo_1,
                'puesto_testigo_2': puesto_testigo_2,
                'puesto_testigo_3': puesto_testigo_3,

            }

            doc.render(context)
            doc.save("/tmp/generated_doc.docx")
            nombre = "CONTRATO_{}.docx".format(contrato)
            f = open('/tmp/generated_doc.docx', mode="rb")
            return http.request.make_response(f.read(),
                                              [('Content-Type', 'application/octet-stream'),
                                               ('Content-Disposition',
                                                'attachment; filename="{}"'.format(nombre))
                                               ])
        except Exception as e:
            return "Upss! algo salio mal en: " + str(e)


class Test_(http.Controller):
    @http.route('/test',  type='http', auth='public', website=True)
    def test(self):
           
        doc = DocxTemplate("/usr/lib/python3/dist-packages/odoo/odoo-extra-addons/proceso_contratacion/templates/PERSONA_FISICA_CONTRATO.docx")
        
        
        context = { 'company_name' : "World company" }
        
        doc.render(context)
        
        doc.save("/tmp/generated_doc.docx")
        nombre = "doc.docx"
        f = open('/tmp/generated_doc.docx', mode="rb")
        return http.request.make_response(f.read(),
                                            [('Content-Type', 'application/octet-stream'),
                                            ('Content-Disposition',
                                                'attachment; filename="{}"'.format(nombre))
                                            ])



class Finiquito(http.Controller):
    @http.route('/documento/FINIQUITO/<contrato>/idplantilla/<id>', type='http', auth='public', website=True) #nuevo
    def finiquito(self, contrato, id): #nuevo
        try: 
            busca_platilla = http.request.env['plantillas.plantillas'].sudo().browse(int(id))[0] #nuevo
            xd = binascii.a2b_base64(busca_platilla.subir_documento) #nuevo
            with open("/tmp/doc2.doc", 'wb') as file: #nuevo
                file.write(xd) #nuevo
            doc = DocxTemplate("/tmp/doc2.doc") #nuevo
            
            finiquito = http.request.env['partidas.partidas'].sudo().browse(int(contrato))[0]

            for f in finiquito:
                numero_contrato = f.nombre_contrato
                fecha_finiquito_ = f.fecha.strftime("%d/%m/%Y")
                fecha_contrato = f.fecha
                nombre_partida = f.nombre_partida
                contratista = f.contratista
                representante_contratista = f.contratista.nombre_representante
                calle_contratista = f.contratista.calles
                colonia_contratista = f.contratista.colonia
                numero_contratista = f.contratista.numero
                estado_contratista = f.contratista.estado_entidad.estado
                municipio_contratista = f.contratista.municipio_delegacion
                codigo_postal_contratista = f.contratista.cp

                importe_contratado = f.total # MONTO DE CONTRATO MAS CONVENIOS SI APLICA
                importe_contratado_civa = (f.total * 0.16) + f.total
                monto_contrato_originalsiva = f.monto_partida # MONTO CONTRATO ORIGINAL SIN MODIFICACIONES
                #fecha_termino_trabajos = f.fecha1.strftime("%d/%m/%Y")
                #fyh_terminacion_trabajos = f.fecha2.strftime("%d/%m/%Y %H:%M:%S")
                numero_bitacora_c = f.numero
                nota_bitacora_ter = f.nota1
                #fpro_acta_recep_traba = f.fecha7.strftime("%d/%m/%Y %H:%M:%S")
                #fecha_entrega_obra = f.fecha8.strftime("%d/%m/%Y %H:%M:%S")
                #fecha_finiquito = f.fecha9.strftime("%d/%m/%Y %H:%M:%S")
                #fecha_cierre_advo = f.fecha10.strftime("%d/%m/%Y %H:%M:%S")
                #fecha_nota_bitacora = f.fecha3.strftime("%d/%m/%Y")
                #fecha_aviso_term_trab = f.fecha4.strftime("%d/%m/%Y")
                #fecha_inicio_real_contrato = f.fecha5.strftime("%d/%m/%Y")
                #fecha_termino_real_contrato = f.fecha6.strftime("%d/%m/%Y")
                #fecha_extincion_derechos = f.fecha11.strftime("%d/%m/%Y %H:%M:%S")
                descripcion_trabajos = f.obra.descripcion
                nombre_obra = f.obra.descripcion
                creditos_cc_finalizar_obra = f.creditosContra
                obra = f.obra
                total_estimacion = f.total_estimacion
                importe_convenio_modificatorio = f.convenios_
                saldo_cantidad = total_estimacion - importe_contratado
                total_amort_anticipo = f.total_amort_anticipo
            
                lista_ruta = []
                for x in f.ruta_critica:
                    if str(x.frente.nombre) == 'False':
                        pass
                    else:
                        lista_ruta.append({'frentes':x.frente.nombre})

                ahora = datetime.now()
                meses = ["Unknown",
                        "Enero",
                        "Febrero",
                        "Marzo",
                        "Abril",
                        "Mayo",
                        "Junio",
                        "Julio",
                        "Agosto",
                        "Septiembre",
                        "Octubre",
                        "Noviembre",
                        "Diciembre"]

                nombre_dias = [
                    "Unknown",
                    "Lunes",
                    "Martes",
                    "Miercoles",
                    "Jueves",
                    "Viernes",
                    "Sabado",
                    "Domingo"]

                nd = ahora.strftime("%A")
                if nd == 'Monday':
                    nd = nombre_dias[1]
                elif nd == 'Tuesday':
                    nd = nombre_dias[2]
                elif nd == 'Wednesday':
                    nd = nombre_dias[3]
                elif nd == 'Thursday':
                    nd = nombre_dias[4]
                elif nd == 'Friday':
                    nd = nombre_dias[5]
                elif nd == 'Saturday':
                    nd = nombre_dias[6]
                elif nd == 'Sunday':
                    nd = nombre_dias[7]

                fecha = str(nd) + " " + str(ahora.day) + " de " + meses[ahora.month] + " de " + str(ahora.year)
                fecha_contrato_letra = str(nd) + " " + str(fecha_contrato.day) + " de " + meses[fecha_contrato.month] + " de " + str(fecha_contrato.year)
                dia = ahora.day
                mes = ahora.month
                year = ahora.year

                numero_fianza = ''
                afianzadora = ''
                fecha_fianza = ''
                monto_fianza = ''
                for fianza in f.numero_contrato.fianzas:
                    numero_fianza = fianza.numero_fianza_fianzas
                    afianzadora = fianza.afianzadora_fianzas
                    fecha_fianza = fianza.fecha_fianza_fianzas
                    monto_fianza = float(fianza.monto)

                
                fecha_inicio_contractual = f.fechainicio
                fletra_inicio_contractual = str(fecha_inicio_contractual.day) + " de " + meses[fecha_inicio_contractual.month] + " de " + str(fecha_inicio_contractual.year)
                fecha_termino_contractual = f.fechatermino
                fletra_termino_contractual = str(fecha_termino_contractual.day) + " de " + meses[fecha_termino_contractual.month] + " de " + str(fecha_termino_contractual.year)
                acum_cont = 0
                lista = []
                lista_convenios = []
                fecha_inicio_conv = ''
                fecha_termino_conv = ''
                periodo = ''
                acum_conv = 0
                for conv_periodo in f.convenio_semaforo:
                    
                    if conv_periodo.plazo_fecha_inicio:
                        fecha_inicio_convenioletra = str(conv_periodo.plazo_fecha_inicio.day) + " de " + meses[conv_periodo.plazo_fecha_inicio.month] + " de " + str(conv_periodo.plazo_fecha_inicio.year)
                        fecha_termino_convenioletra = str(conv_periodo.plazo_fecha_termino.day) + " de " + meses[conv_periodo.plazo_fecha_termino.month] + " de " + str(conv_periodo.plazo_fecha_termino.year)
                        acum_conv += 1
                        periodo =  str(fecha_inicio_convenioletra) + ' al ' + str(fecha_termino_convenioletra)
                    elif acum_conv == 0:
                        fecha_inicio_convenioletra = str(f.fechainicio.day) + " de " + meses[f.fechainicio.month] + " de " + str(f.fechainicio.year)
                        fecha_termino_convenioletra = str(f.fechatermino.day) + " de " + meses[f.fechatermino.month] + " de " + str(f.fechatermino.year)
                        periodo = str(fecha_inicio_convenioletra) + ' al ' + str(fecha_termino_convenioletra) 

                for conv in f.convenio_semaforo:
                    acum_cont += 1
                    '''if f.fecha_inicio_convenida:
                        periodo =  str(f.fecha_inicio_convenida) + ' al ' + str(f.fecha_termino_convenida)
                    else:
                        periodo = str(f.fechainicio) + ' al ' + str(f.fechatermino)'''
                    if acum_cont == 1:
                        periodo = periodo
                    else:
                        periodo = ''

                    if conv.tipo_convenio != 'OB':
                        if conv.tipo_convenio == 'MT':
                            if fecha_inicio_conv == '': 
                                fecha_inicio_conv = fecha_inicio_contractual
                                fecha_termino_conv = fecha_termino_contractual
                            lista_convenios.append({'documento': str(numero_contrato) + '-' + str(conv.referencia), 'fecha_inicio': fecha_inicio_conv, 'fecha_termino': fecha_termino_conv, 'periodo': periodo})
                        else:
                            fecha_inicio_conv = conv.plazo_fecha_inicio
                            fecha_termino_conv = conv.plazo_fecha_termino
                            lista_convenios.append({'documento': str(numero_contrato) + '-' + str(conv.referencia), 'fecha_inicio': fecha_inicio_conv, 'fecha_termino': fecha_termino_conv, 'periodo': periodo})

                    
                    if acum_cont == 1:
                        lista.append({'label': 'CONTRATO', 'cols': fletra_inicio_contractual, 'cols2': fletra_termino_contractual, 'cols3': periodo})
                    else:
                        nombre_convenio = 'CONVENIO ' + str(conv.referencia) + '\n' + str(conv.tipo_convenio)
                        lista.append({'label': nombre_convenio, 'cols': conv.plazo_fecha_inicio, 'cols2': conv.plazo_fecha_termino, 'cols3': '',})
                

                lista_esti = []
                
                for est in f.esti:
                    if est.tipo_estimacion == '2':
                        escalatoria = est.idobra
                    else:
                        escalatoria = ''

                    periodo_est =  str(est.fecha_inicio_estimacion) + ' al ' + str(est.fecha_termino_estimacion)
                    lista_esti.append({'label': est.idobra, 'cols': escalatoria, 'cols2': str(est.fecha_presentacion.day) + meses[est.fecha_presentacion.month] + str(est.fecha_presentacion.year), 
                    'cols3': periodo_est, 'cols4': '{:20,.2f}'.format(est.estimado), 
                    'cols5': '{:20,.2f}'.format(float(est.amort_anticipo)),})

            titular = http.request.env['ir.config_parameter'].sudo().get_param('firmas_logos.nombre_titular')
            contratistas = http.request.env['contratista.contratista'].sudo().browse(int(contratista))[0]

            for c in contratistas:
                contratista = c.name

            obras = http.request.env['registro.programarobra'].sudo().browse(int(obra))[0]
            for o in obras:
                descripcion_obra = o.descripcion

            context = {
                'fecha_completa': fecha,
                'dia': dia,
                'mes': mes,
                'year': year,
                'numero_contrato': numero_contrato,
                'fecha_contrato_letra': fecha_contrato_letra,
                #'fecha_finiquito_': fecha_finiquito,
                'nombre_partida': nombre_partida,
                'contratista': contratista,

                'calle_contratista': calle_contratista , 
                'colonia_contratista':  colonia_contratista, 
                'numero_contratista': numero_contratista , 
                'estado_contratista':  estado_contratista,  
                'municipio_contratista':  municipio_contratista,  
                'codigo_postal_contratista':  codigo_postal_contratista, 

                'importe_contratado': '{:20,.2f}'.format(float(importe_contratado)),
                'importe_contratado_civa': '{:20,.2f}'.format(float(importe_contratado_civa)), 
                'monto_contrato_originalsiva': '{:20,.2f}'.format(float(monto_contrato_originalsiva)),
                'total_estimacion': '{:20,.2f}'.format(float(total_estimacion)),
                'importe_convenio_modificatorio': '{:20,.2f}'.format(float(importe_convenio_modificatorio)),
                'saldo_cantidad': '{:20,.2f}'.format(float(saldo_cantidad)),
                'total_amort_anticipo': '{:20,.2f}'.format(float(total_amort_anticipo)),

                #'fecha_termino_trabajos': fecha_termino_trabajos,
                #'fyh_terminacion_trabajos': fyh_terminacion_trabajos,
                'numero_bitacora_c': numero_bitacora_c,
                'nota_bitacora_ter': nota_bitacora_ter,
                #'fpro_acta_recep_traba': fpro_acta_recep_traba,
                #'fecha_entrega_obra': fecha_entrega_obra,
                #'fecha_finiquito': fecha_finiquito,
                #'fecha_cierre_advo': fecha_cierre_advo,
                #'fecha_nota_bitacora': fecha_nota_bitacora,
                #'fecha_aviso_term_trab': fecha_aviso_term_trab,
                #'fecha_inicio_real_contrato': fecha_inicio_real_contrato,
                #'fecha_termino_real_contrato': fecha_termino_real_contrato,
                #'fecha_extincion_derechos': fecha_extincion_derechos,
                'descripcion_trabajos': descripcion_trabajos,
                'creditos_cc_finalizar_obra': creditos_cc_finalizar_obra,
                'descripcion_obra': descripcion_obra, 
                'nombre_obra': nombre_obra,
                'representante_contratista': representante_contratista,
                'titular': titular, # TITULAR DE SIDUR
                'numero_fianza': numero_fianza,
                'afianzadora': afianzadora,
                'fecha_fianza': fecha_fianza, 
                'monto_fianza': monto_fianza,
                # TABLA
                'tbl_contents': lista,
                'tbl_estimacion': lista_esti,
                'frentes':lista_ruta,
                'lista_convenios': lista_convenios, 

            }    

            print(context)
            doc.render(context)

            doc.save("/tmp/generated_doc.docx")
            nombre = "documento_{}.docx".format(contrato)
            f = open('/tmp/generated_doc.docx', mode="rb")
            return http.request.make_response(f.read(),
                                            [('Content-Type', 'application/octet-stream'),
                                            ('Content-Disposition',
                                                'attachment; filename="{}"'.format(nombre))
                                            ])
        except Exception as e:
            return "Upss! algo salio mal. Error en: " + str(e)



class LicitacionConvocatoriaPublica(http.Controller):
    @http.route('/documento/LICITACION/<numerolicitacion>/idplantilla/<id>', type='http', auth='public', website=True) #nuevo
    def licitacion(self, numerolicitacion, id): #nuevo
        try: 
            busca_platilla = http.request.env['plantillas.plantillas'].sudo().browse(int(id))[0] #nuevo

            xd = binascii.a2b_base64(busca_platilla.subir_documento) #nuevo
            with open("/tmp/Documento.DOCX", 'wb') as file: #nuevo
                file.write(xd) #nuevo
            
            doc = DocxTemplate("/tmp/Documento.DOCX") #nuevo

            licitacion = http.request.env['proceso.licitacion'].sudo().browse(int(id))[0]

            for i in licitacion:
                b_lic = http.request.env['proceso.licitacion'].browse(i.id)
                for x in b_lic:
                    for rec in x.programar_obra_licitacion:
                        oficio = rec.recursos.name.name
                        fecha_oficio_ = rec.recursos.name.fecha_actual

                    numero_licitacion = x.numerolicitacion
                    ahora = datetime.now()
                    meses = ["Unknown",
                            "Enero",
                            "Febrero",
                            "Marzo",
                            "Abril",
                            "Mayo",
                            "Junio",
                            "Julio",
                            "Agosto",
                            "Septiembre",
                            "Octubre",
                            "Noviembre",
                            "Diciembre"]
                    nombre_dias = [
                        "Unknown",
                        "Lunes",
                        "Martes",
                        "Miercoles",
                        "Jueves",
                        "Viernes",
                        "Sabado",
                        "Domingo"]
                    nd = ahora.strftime("%A")
                    if nd == 'Monday':
                        nd = nombre_dias[1]
                    elif nd == 'Tuesday':
                        nd = nombre_dias[2]
                    elif nd == 'Wednesday':
                        nd = nombre_dias[3]
                    elif nd == 'Thursday':
                        nd = nombre_dias[4]
                    elif nd == 'Friday':
                        nd = nombre_dias[5]
                    elif nd == 'Saturday':
                        nd = nombre_dias[6]
                    elif nd == 'Sunday':
                        nd = nombre_dias[7]

                    dia_fecha_con_inv = x.fechaconinv.strftime("%A")
                    if dia_fecha_con_inv == 'Monday':
                        dia_fecha_con_inv = nombre_dias[1]
                    elif dia_fecha_con_inv == 'Tuesday':
                        dia_fecha_con_inv = nombre_dias[2]
                    elif dia_fecha_con_inv == 'Wednesday':
                        dia_fecha_con_inv = nombre_dias[3]
                    elif dia_fecha_con_inv == 'Thursday':
                        dia_fecha_con_inv = nombre_dias[4]
                    elif dia_fecha_con_inv == 'Friday':
                        dia_fecha_con_inv = nombre_dias[5]
                    elif dia_fecha_con_inv == 'Saturday':
                        dia_fecha_con_inv = nombre_dias[6]
                    elif dia_fecha_con_inv == 'Sunday':
                        dia_fecha_con_inv = nombre_dias[7]

                    visita = x.visitafechahora.strftime("%A")
                    if visita == 'Monday':
                        visita = nombre_dias[1]
                    elif visita == 'Tuesday':
                        visita = nombre_dias[2]
                    elif visita == 'Wednesday':
                        visita = nombre_dias[3]
                    elif visita == 'Thursday':
                        visita = nombre_dias[4]
                    elif visita == 'Friday':
                        visita = nombre_dias[5]
                    elif visita == 'Saturday':
                        visita = nombre_dias[6]
                    elif visita == 'Sunday':
                        visita = nombre_dias[7]

                    fecha_visita = str(x.visitafechahora.day) + " de " + meses[x.visitafechahora.month] \
                                + " del " + str(x.visitafechahora.year) + ' a las ' + str(x.visitafechahora.hour) + ':' + str(x.visitafechahora.minute) + '0 Horas'

                    apertura = x.aperturafechahora.strftime("%A")
                    if apertura == 'Monday':
                        apertura = nombre_dias[1]
                    elif apertura == 'Tuesday':
                        apertura = nombre_dias[2]
                    elif apertura == 'Wednesday':
                        apertura = nombre_dias[3]
                    elif apertura == 'Thursday':
                        apertura = nombre_dias[4]
                    elif apertura == 'Friday':
                        apertura = nombre_dias[5]
                    elif apertura == 'Saturday':
                        apertura = nombre_dias[6]
                    elif apertura == 'Sunday':
                        apertura = nombre_dias[7]

                    fecha_apertura = str(x.aperturafechahora.day) + " de " + meses[x.aperturafechahora.month] \
                                + " del " + str(x.aperturafechahora.year) + ' a las ' + str(x.aperturafechahora.hour) + ':' + str(x.aperturafechahora.minute) + '0 Horas'

                    junta = x.juntafechahora.strftime("%A")
                    if junta == 'Monday':
                        junta = nombre_dias[1]
                    elif junta == 'Tuesday':
                        junta = nombre_dias[2]
                    elif junta == 'Wednesday':
                        junta = nombre_dias[3]
                    elif junta == 'Thursday':
                        junta = nombre_dias[4]
                    elif junta == 'Friday':
                        junta = nombre_dias[5]
                    elif junta == 'Saturday':
                        junta = nombre_dias[6]
                    elif junta == 'Sunday':
                        junta = nombre_dias[7]

                    fallo = x.fallofechahora.strftime("%A")
                    if fallo == 'Monday':
                        fallo = nombre_dias[1]
                    elif fallo == 'Tuesday':
                        fallo = nombre_dias[2]
                    elif fallo == 'Wednesday':
                        fallo = nombre_dias[3]
                    elif fallo == 'Thursday':
                        fallo = nombre_dias[4]
                    elif fallo == 'Friday':
                        fallo = nombre_dias[5]
                    elif fallo == 'Saturday':
                        fallo = nombre_dias[6]
                    elif fallo == 'Sunday':
                        fallo = nombre_dias[7]

                    fecha_fallo = str(fallo) + " " + str(x.fallofechahora.day) + " de " + meses[x.fallofechahora.month] \
                                    + " del " + str(x.fallofechahora.year)
                    fallo_hora = str(x.fallofechahora.hour) + ':' + str(x.fallofechahora.minute) + '0 Horas'

                    fecha_junta = str(x.juntafechahora.day) + " de " + meses[x.juntafechahora.month] + " del " + \
                                str(x.juntafechahora.year) + ' a las ' + str(x.juntafechahora.hour) + ':' + str(x.juntafechahora.minute) + '0 Horas'

                    fecha = str(nd) + " " + str(ahora.day) + " de " + meses[ahora.month] + " del " + str(ahora.year)

                    fecha_con_inv = str(x.fechaconinv.day) + " de " + meses[x.fechaconinv.month] + " del " + str(x.fechaconinv.year)
                    fecha_oficio = str(fecha_oficio_.day) + " de " + meses[fecha_oficio_.month] + " del " + str(fecha_oficio_.year)
                    fecha_inicio_licitacion = str(x.fechaestimadainicio.day) + " de " + meses[x.fechaestimadainicio.month] + " del " + str(x.fechaestimadainicio.year)
                    fecha_termino_licitacion = str(x.fechaestimadatermino.day) + " de " + meses[x.fechaestimadatermino.month] + " del " + str(x.fechaestimadatermino.year)
                    fecha_limite_bases = str(x.fechalimiteentregabases.day) + " de " + meses[x.fechalimiteentregabases.month] \
                                        + " del " + str(x.fechalimiteentregabases.year)

                    dia = ahora.day
                    mes = ahora.month
                    mes_letra = meses[ahora.month]
                    year = ahora.year
                    context = {
                        'dia': dia,
                        'mes_letra': mes_letra,
                        'year': year,
                        'fecha_completa': fecha, # FECHA DE HOY
                        'fecha_con_inv': fecha_con_inv, # FECHA CON/INV
                        'fecha_limite_bases': fecha_limite_bases, # FECHA LIMITE PARA LA ENTREGA DE BASES
                        'numero_licitacion': numero_licitacion, # NUMERO DE LA LICITACION
                        'concepto': x.name, # OBJETO DE LA LICITACION
                        'oficio': oficio, # OFICIO DEL RECURSO
                        'fecha_oficio': fecha_oficio, # FECHA DEL OFICIO
                        'plazo_dias': x.plazodias, # PLAZO DE DIAS DE LA LICITACION
                        'fecha_inicio_licitacion': fecha_inicio_licitacion, # FECHA ESTIMADA DE INICIO
                        'fecha_termino_licitacion': fecha_termino_licitacion, # FECHA ESTIMADA DE TERMINO
                        'fecha_visita': fecha_visita, # FECHA Y HORA DE LA VISITA DE OBRA
                        'lugar_visita': x.visitalugar, # LUGAR DE LA VISITA DE OBRA
                        'fecha_junta': fecha_junta,# FECHA Y HORA DE LA JUNTA DE ACLARACIONES
                        'lugar_junta': x.juntalugar, # LUGAR DE LA JUNTA DE ACLARACIONES
                        'fecha_apertura': fecha_apertura, # FECHA Y HORA DE LA APERTURA
                        'lugar_apertura': x.aperturalugar, # LUGAR DE LA APERTURA
                        'capital_contable': '{:20,.2f}'.format(x.capitalcontable), # CAPITAL CONTABLE
                        'dia_fecha_con_inv': dia_fecha_con_inv, # DIA CON LETRA DE LA FECHA CON/INV
                        'convocatoria': x.convocatoria, # CONVOCATORIA

                        'capital_contable_letra': num2words(x.capitalcontable, lang='es'), # CAPITAL CONTABLE CON LETRA
                        'anticipo_inicio': int(x.anticipoinicio), # PORCENTAJE DE ANTICIPO
                        'porcentaje_letra': num2words(x.anticipoinicio, lang='es'), # PORCENTAJE DE ANTICIPO CON LETRA
                        'fecha_fallo': fecha_fallo, # FECHA FALLO
                        'fallo_hora': fallo_hora, # HORA DEL FALLO
                        'lugar_fallo': x.fallolugar, # LUGAR DEL FALLO
                     }

                    doc.render(context)

                    doc.save("/tmp/generated_Documento.docx")
                    nombre = "Documento_{}.docx".format(id)
                    f = open('/tmp/generated_Documento.docx', mode="rb")
                    return http.request.make_response(f.read(),
                                                    [('Content-Type', 'application/octet-stream'),
                                                    ('Content-Disposition',
                                                        'attachment; filename="{}"'.format(nombre))
                                                    ])


        except Exception as e:
            return "Upss! algo salio mal. Error en: " + str(e)


class FichaTecnica(http.Controller):
    @http.route('/FICHA_TECNICA/FICHA_TECNICA/', type='http', auth='public', website=True)
    def ficha(self, **kw):
        try:
            busca_platilla = http.request.env['plantillas.plantillas'].sudo().search([('tipo_documento.name', '=', 'FICHA_TECNICA')]) # nuevo

            xd = binascii.a2b_base64(busca_platilla.subir_documento)  # nuevo
            with open("/tmp/doc2.doc", 'wb') as file:  # nuevo

                file.write(xd)  # nuevo
            doc = DocxTemplate("/tmp/doc2.doc")  # nuevo

            qwerty = http.request.env['partidas.partidas'].sudo().search([('id', '=', kw['id'])])
            for i in qwerty:
                nombre_obra = i.obra.descripcion
                nombre_contrato = i.nombre_contrato  # Número de contrato
                residente = ''

                for res in i.residente_obra:
                    residente = res.name

                municipio_localidad = str(i.obra.obra_planeada.localidad) + ', ' + str(i.obra.obra_planeada.municipio)
                tipo_obra = i.tipo_obra.tipo_obra
                ubicacion = i.obra.obra_planeada.ubicacion
                monto_obra_siva = i.total
                actividades = i.numero_contrato.descripciontrabajos

                fecha_contrato_inicio = ''
                fecha_contrato_termino = ''

                if i.fecha_inicio_convenida: # con convenio
                    fecha_contrato_inicio = i.fecha_inicio_convenida
                    fecha_contrato_termino = i.fecha_termino_convenida
                else: # fecha normal
                    fecha_contrato_inicio = i.fechainicio
                    fecha_contrato_termino = i.fechatermino

                meses = ["Unknown",
                            "Enero",
                            "Febrero",
                            "Marzo",
                            "Abril",
                            "Mayo",
                            "Junio",
                            "Julio",
                            "Agosto",
                            "Septiembre",
                            "Octubre",
                            "Noviembre",
                            "Diciembre"]

                fecha_inicio = str(fecha_contrato_inicio.day) + " de " + meses[fecha_contrato_inicio.month] + " de " + str(fecha_contrato_inicio.year)
                fecha_termino = str(fecha_contrato_termino.day) + " de " + meses[fecha_contrato_termino.month] + " de " + str(fecha_contrato_termino.year)

            context = {
                'nombre_obra': nombre_obra,
                'contrato': nombre_contrato,
                'tipo_obra': tipo_obra,
                'monto_obra_siva': '{:20,.2f}'.format(monto_obra_siva),  # Monto del Contrato Sin IVA
                'residente': residente,
                'municipio_localidad': municipio_localidad,
                'ubicacion': ubicacion,
                'fecha_contrato_inicio': fecha_contrato_inicio,
                'fecha_contrato_termino': fecha_contrato_termino,
                'actividades': actividades,

            }

            doc.render(context)
            doc.save("/tmp/generated_doc.docx")
            nombre = "Ficha_Tecnica_{}.docx".format(nombre_contrato)
            f = open('/tmp/generated_doc.docx', mode="rb")
            return http.request.make_response(f.read(),
                                                [('Content-Type', 'application/octet-stream'),
                                                ('Content-Disposition',
                                                'attachment; filename="{}"'.format(nombre))
                                                ])
        except Exception as e:
            return "Upss! algo salio mal en: " + str(e)


class DescargarConceptos(http.Controller):
    @http.route('/Conceptos/', auth='public')
    def index(self, **kw):
        try:
            partidas = http.request.env['partidas.partidas'].search(
                [('id', '=', kw['id'])])  # , ('idobra', '=', '1')
            workbook = load_workbook(
                filename="/usr/lib/python3/dist-packages/odoo/odoo-extra-addons/proceso_contratacion/static/conceptos_pantilla.xlsx")
            sheet = workbook.active
            wb = Workbook()

            fill = PatternFill(fill_type=None, start_color='bdbdbd', end_color='bdbdbd')
            double = Side(border_style="double", color="000000")
            thin = Side(border_style="thin", color="000000")

            nombre_partida = ''
            contratista = ''

            for o in partidas:
                b_partida = http.request.env['partidas.partidas'].browse(o.id)

                nombre_partida = b_partida.nombre_partida
                contratista = b_partida.contratista.name

                acum = 0
                pos_concepto = 0
                aviso = 0
                xd = 0
                for y in b_partida.conceptos_partidas:
                    pos_concepto += 1
                    if str(y.medida) == 'False':
                        y.medida = ''
                    if str(y.cantidad) == '0.0':
                        y.cantidad = ''
                    if str(y.precio_unitario) == '0.0':
                        y.precio_unitario = ''
                    if str(y.importe) == '0.0':
                        y.importe = ''
                    # AVISO INDICA CUANDO HAY QUE APLICAR SUBTOTAL
                    # CLAVE

                    acum += 1 + aviso
                    # CLAVE
                    for column in range(2, 3):
                        column_letter = get_column_letter(column)
                        xd = acum + 11
                        aviso = 0
                        sheet[column_letter + str(xd)] = str(y.clave_linea)
                        sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin,
                                                                       bottom=thin)
                    # CONCEPTO
                    for column in range(3, 4):
                        column_letter = get_column_letter(column)
                        xd = acum + 11
                        aviso = 0
                        sheet[column_letter + str(xd)] = str(y.concepto)
                        sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin,
                                                                       bottom=thin)
                    # MEDIDA
                    for column in range(4, 5):
                        column_letter = get_column_letter(column)
                        xd = acum + 11
                        aviso = 0
                        sheet[column_letter + str(xd)] = str(y.medida)
                        sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin,
                                                                       bottom=thin)
                    # CANTIDAD
                    for column in range(5, 6):
                        column_letter = get_column_letter(column)
                        xd = acum + 11
                        aviso = 0
                        sheet[column_letter + str(xd)] = str(y.cantidad)
                        sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin,
                                                                       bottom=thin)
                    # PRECIO UNITARIO
                    for column in range(6, 7):
                        column_letter = get_column_letter(column)
                        xd = acum + 11
                        aviso = 0
                        sheet[column_letter + str(xd)] = y.precio_unitario
                        sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin,
                                                                       bottom=thin)
                        sheet[column_letter + str(xd)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                    # IMPORTE
                    for column in range(7, 8):
                        column_letter = get_column_letter(column)
                        xd = acum + 11
                        aviso = 0
                        sheet[column_letter + str(xd)] = y.importe
                        sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin,
                                                                       bottom=thin)
                        sheet[column_letter + str(xd)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'

                column_letteru = get_column_letter(6)  # TOTAL
                column_letterx = get_column_letter(7)  # TOTAL
                sheet[column_letteru + str(xd + 1)] = 'TOTAL'
                sheet[column_letteru + str(xd + 1)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                sheet[column_letteru + str(xd + 1)].fill = PatternFill("solid", fgColor="ff7043")
                sheet[column_letteru + str(xd + 1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                sheet[column_letterx + str(xd + 1)] = "=SUM(" + column_letterx + str(12) + ":" + column_letterx + str(
                    xd) + ")"
                sheet[column_letterx + str(xd + 1)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                sheet[column_letterx + str(xd + 1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            for column in range(6, 7):
                column_letter = get_column_letter(column)
                sheet[column_letter + "8"] = nombre_partida

            for column in range(3, 4):
                column_letter = get_column_letter(column)
                sheet[column_letter + "8"] = contratista

            # Save the spreadsheet
            workbook.save("/tmp/conceptos.xlsx")

            # prs.save('/tmp/test.pptx')
            f = open('/tmp/conceptos.xlsx', mode="rb")
            nombre = "Conceptos_{}.docx".format(nombre_partida)
            return http.request.make_response(f.read(),
                                              [('Content-Type', 'application/octet-stream'),
                                               ('Content-Disposition',
                                                'attachment; filename="{}"'.format(nombre,'conceptos.xlsx'))
                                               ])

        except Exception as e:
            return "Upss! algo salio mal en: " + str(e)