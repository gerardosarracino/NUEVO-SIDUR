# -*- coding: utf-8 -*-
from odoo import http

from datetime import datetime, date, time, timedelta
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from openpyxl.styles import colors
from openpyxl.styles import Font, Color, NamedStyle
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment


class FiniquitoExcel(http.Controller):
    @http.route('/finiquito_excel/finiquito_excel/', auth='public')
    def index(self, **kw):
        try:
            # workbook = load_workbook(filename="/home/gerardo/Developments/odoo12/extra-addons/finiquito_excel/static/plantilla.xlsx")
            workbook = load_workbook(filename="/usr/lib/python3/dist-packages/odoo/odoo-extra-addons/finiquito_excel/static/plantilla.xlsx")
            sheet = workbook.active
            wb = Workbook()

            # FASTES METHOD
            # wb = Workbook()
            # ws = wb.new_sheet("FINIQUITO")

            nombre_partida = ''
            contratista = ''
            total_contrato = ''
            # fill = PatternFill(fill_type=None, start_color='bdbdbd', end_color='bdbdbd')
            # double = Side(border_style="double", color="000000")
            # thin = Side(border_style="thin", color="000000")

            '''bordes = NamedStyle(name="bordes")
            bordes.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            wb.add_named_style(bordes)'''

            # ESTILOS DE SUBTOTAL
            subtotal_estilo = NamedStyle(name="subtotal_estilo")
            subtotal_estilo.font = Font(bold=True, size=9)
            subtotal_estilo.fill = PatternFill("solid", fgColor="ff7043")
            # subtotal_estilo.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            wb.add_named_style(subtotal_estilo)

            # ESTILOS DE CATEGORIAS
            categoria_estilo = NamedStyle(name="categoria_estilo")
            # categoria_estilo.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            categoria_estilo.font = Font(bold=True, size=9)
            categoria_estilo.fill = PatternFill("solid", fgColor="bdbdbd")
            wb.add_named_style(categoria_estilo)

            # ESTILOS DE HEADER DE ESTIMACIONES
            header_est_estilo = NamedStyle(name="header_est_estilo")
            header_est_estilo.font = Font(name='Arial', size=9, bold=True)
            header_est_estilo.fill = PatternFill("solid", fgColor="ffff00")
            # header_est_estilo.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            wb.add_named_style(header_est_estilo)

            # ESTILOS TOTAL
            total_estilo = NamedStyle(name="total_estilo")
            # total_estilo.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            total_estilo.fill = PatternFill("solid", fgColor="e65100")
            total_estilo.font = Font(color="fafafa", name='Arial', size=9, bold=True)
            wb.add_named_style(total_estilo)

            http.request.env.cr.execute("SELECT clave_linea,concepto,medida,cantidad,precio_unitario,importe FROM proceso_conceptos_part WHERE id_partida = " + kw['id'] + " ORDER BY id ASC")
            res_concepto = http.request.cr.fetchall()

            # COLUMNA DE TOTALES
            estimacion = http.request.env['control.estimaciones']
            estimacion_count = estimacion.search_count([('obra.id', '=', kw['id'])])
            columna_total_cantidad = get_column_letter(8 + (estimacion_count * 2))
            columna_total_importe = get_column_letter(9 + (estimacion_count * 2))

            acum_conceptos = 10  # INICIA A PARTIR DE LA FILA # 10 Y SE VA ACUMULANDO PARA IR AGREGANDO UN ESPACIO ABAJO CADA OBJETO
            for i in res_concepto:
                acum_conceptos += 1
                columna_clave = get_column_letter(2)  # COLUMNA CLAVE
                columna_concepto = get_column_letter(3)  # COLUMNA CONCEPTO
                columna_unidad = get_column_letter(4)  # COLUMNA UNIDAD
                columna_cantidad = get_column_letter(5)  # COLUMNA CANTIDAD
                columna_precio = get_column_letter(6)  # COLUMNA PRECIO UNITARIO
                columna_importe = get_column_letter(7)  # COLUMNA IMPORTE
                sheet[columna_clave + str(acum_conceptos)] = i[0]  # AGREGA LA CLAVE EN SU POSICION
                # sheet[columna_clave + str(acum_conceptos)].style = bordes
                sheet[columna_concepto + str(acum_conceptos)] = i[1]  # AGREGA CONCETP EN SU POSICION
                # sheet[columna_concepto + str(acum_conceptos)].style = bordes
                sheet[columna_unidad + str(acum_conceptos)] = i[2]  # AGREGA UNIDAD EN SU POSICION
                sheet[columna_cantidad + str(acum_conceptos)] = i[3]  # AGREGA CANTIDAD EN SU POSICION
                sheet[columna_precio + str(acum_conceptos)] = i[4]  # AGREGA PRECIO U EN SU POSICION
                sheet[columna_precio + str(acum_conceptos)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                sheet[columna_importe + str(acum_conceptos)] = i[5]  # AGREGA IMPORTE EN SU POSICION
                sheet[columna_importe + str(acum_conceptos)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'

                if not i[4]:  # ES CATEGORIA PASAR
                    sheet[columna_clave + str(acum_conceptos)].style = categoria_estilo
                    sheet[columna_concepto + str(acum_conceptos)].style = categoria_estilo
                    sheet[columna_unidad + str(acum_conceptos)].style = categoria_estilo
                    sheet[columna_cantidad + str(acum_conceptos)].style = categoria_estilo
                    sheet[columna_precio + str(acum_conceptos)].style = categoria_estilo
                    sheet[columna_importe + str(acum_conceptos)].style = categoria_estilo
                    sheet[columna_total_cantidad + str(acum_conceptos)].style = total_estilo
                    sheet[columna_total_importe + str(acum_conceptos)].style = total_estilo
                else:
                    # TOTALES T.CANTIDAD
                    http.request.env.cr.execute(
                        "SELECT COALESCE(SUM(estimacion),0),c FROM control_detalle_conceptos WHERE id_partida = " +
                        kw['id'] + " AND estimacion != 0 AND clave_linea = '" + str(
                            i[0]) + "' AND precio_unitario = " + str(i[4]) + " AND cantidad = " + str(i[3]))
                    res_importe = http.request.cr.fetchall()
                    for t in res_importe:
                        sheet[columna_total_cantidad + str(acum_conceptos)] = t[0]
                        sheet[columna_total_cantidad + str(acum_conceptos)].style = total_estilo
                        sheet[columna_total_importe + str(acum_conceptos)] = t[1]
                        sheet[columna_total_importe + str(acum_conceptos)].style = total_estilo
                        sheet[columna_total_importe + str(acum_conceptos)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'

                    http.request.env.cr.execute("SELECT estimacion,clave_linea,precio_unitario,importe_ejecutado,num_est,cantidad FROM control_detalle_conceptos WHERE id_partida = " + kw['id'] + " AND estimacion != 0 AND clave_linea = '" + str(i[0]) + "' AND precio_unitario = " + str(i[4]) + " AND cantidad = " + str(i[3]))
                    res = http.request.cr.fetchall()

                    columna_cantidad_estimacion = get_column_letter(8)  # COLUMNA DE CANTIDAD DE ESTIMACION
                    columna_importe_estimacion = get_column_letter(9)  # COLUMNA DE IMPORTE DE ESTIMACION

                    # sheet[columna_cantidad_estimacion + str(acum_conceptos)].style = bordes
                    # sheet[columna_importe_estimacion + str(acum_conceptos)].style = bordes
                    for q in res:
                        print('CLAVE DE LINEA:',q[1], '-------  ESTIMACION:', q[4])
                        if int(q[4]) == 1: # q[4] = numero de estimacion
                            sheet[columna_cantidad_estimacion + str(acum_conceptos)] = q[0]  # AGREGA LA CANTIDAD EST
                            sheet[columna_importe_estimacion + str(acum_conceptos)] = q[3]  # AGREGA EL IMPORTE EST
                            sheet[columna_importe_estimacion + str(acum_conceptos)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                        else:
                            columna_cantidad_estimacion = get_column_letter(7 + q[4] + (q[4]-1))  # COLUMNA DE CANTIDAD DE ESTIMACION
                            columna_importe_estimacion = get_column_letter(8 + q[4] + (q[4]-1))  # COLUMNA DE IMPORTE DE ESTIMACION
                            sheet[columna_cantidad_estimacion + str(acum_conceptos)] = q[0]  # AGREGA LA CANTIDAD EST
                            sheet[columna_importe_estimacion + str(acum_conceptos)] = q[3]  # AGREGA EL IMPORTE EST
                            sheet[columna_importe_estimacion + str(acum_conceptos)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'

            # SE AGREGAN LAS COLUMNAS DE ESTIMACION #
            numero_est = 1
            acum_estx = 0
            for r in range(estimacion_count):
                numero_est += 1
                if numero_est == 1:
                    pass
                else:
                    acum_estx += 1
                    columna_numest = get_column_letter(7 + numero_est + acum_estx)
                    columna_cantidadest = get_column_letter(7 + numero_est + acum_estx)
                    columna_importeest = get_column_letter(8 + numero_est + acum_estx)
                    sheet[columna_numest + '9'] = 'ESTIMACION ' + str(numero_est)
                    sheet[columna_numest + '9'].style = header_est_estilo
                    sheet[columna_numest + '9'].alignment = Alignment(horizontal="center", vertical="center")
                    sheet.merge_cells("" + columna_numest + "9:" + columna_importeest + "9")
                    sheet[columna_cantidadest + '10'] = 'CANTIDAD'
                    sheet[columna_cantidadest + '10'].style = header_est_estilo
                    sheet[columna_importeest + '10'] = 'IMPORTE'
                    sheet[columna_importeest + '10'].style = header_est_estilo
            print(acum_conceptos, ' xxxxxxxxxxx ')
            # TOTALES
            http.request.env.cr.execute(
                "SELECT COALESCE(SUM(importe_ejecutado),0) FROM control_detalle_conceptos WHERE id_partida = " + kw['id'])
            res_total = http.request.cr.fetchall()
            for tt in res_total:
                sheet[columna_total_importe + str(acum_conceptos+1)] = tt[0]
                sheet[columna_total_importe + str(acum_conceptos+1)].style = total_estilo
                sheet[columna_total_importe + str(acum_conceptos+1)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'

            sheet[columna_total_cantidad + '9'] = 'TOTALES'
            sheet.merge_cells("" + columna_total_cantidad + "9:" + columna_total_importe + "9")
            sheet[columna_total_cantidad + '9'].style = total_estilo
            sheet[columna_total_cantidad + '9'].alignment = Alignment(wrap_text=True)
            sheet[columna_total_cantidad + '10'] = 'T.CANTIDAD'
            sheet[columna_total_cantidad + '10'].style = total_estilo
            sheet[columna_total_importe + '10'] = 'T.IMPORTE'
            sheet[columna_total_importe + '10'].style = total_estilo
            sheet[columna_total_cantidad + str(acum_conceptos+1)] = 'GRAN TOTAL'
            sheet[columna_total_cantidad + str(acum_conceptos+1)].style = total_estilo
            # sheet[columna_total_importe + str(acum_conceptos+1)].style = total_estilo

            # Save the spreadsheet
            workbook.save("/tmp/finiquito.xlsx")

            '''http.request.env.cr.execute(
                "SELECT clave_linea,concepto,medida,cantidad,precio_unitario,importe FROM proceso_conceptos_part WHERE id_partida = " +
                kw['id'] + " ORDER BY id ASC")
            res_concepto = http.request.cr.fetchall()
            estimacion = http.request.env['control.estimaciones']
            estimacion_count = estimacion.search_count([('obra.id', '=', kw['id'])])
            acum_conceptos = 10  # INICIA A PARTIR DE LA FILA # 10 Y SE VA ACUMULANDO PARA IR AGREGANDO UN ESPACIO ABAJO CADA OBJETO
            for i in res_concepto:
                acum_conceptos += 1
                ws[acum_conceptos][2].value = i[0]  # AGREGA LA CLAVE EN SU POSICION
                ws[acum_conceptos][3].value = i[1]  # AGREGA CONCETP EN SU POSICION
                ws[acum_conceptos][4].value = i[2]  # AGREGA UNIDAD EN SU POSICION
                ws[acum_conceptos][5].value = i[3]  # AGREGA CANTIDAD EN SU POSICION
                ws[acum_conceptos][6].value = i[4]  # AGREGA PRECIO U EN SU POSICION
                ws[acum_conceptos][7].value = i[5]  # AGREGA IMPORTE EN SU POSICION

                if not i[4]:  # ES CATEGORIA PASAR
                    pass
                else:
                    # TOTALES T.CANTIDAD
                    ws[acum_conceptos][8 + (estimacion_count * 2)].value = 'x'
                    ws[acum_conceptos][9 + (estimacion_count * 2)].value = 'x'

                    http.request.env.cr.execute(
                        "SELECT estimacion,clave_linea,precio_unitario,importe_ejecutado,num_est,cantidad FROM control_detalle_conceptos WHERE id_partida = " +
                        kw['id'] + " AND estimacion != 0 AND clave_linea = '" + str(
                            i[0]) + "' AND precio_unitario = " + str(i[4]) + " AND cantidad = " + str(i[3]))
                    res = http.request.cr.fetchall()
                    for q in res:
                        print('CLAVE DE LINEA:', q[1], '-------  ESTIMACION:', q[4])

                        if int(q[4]) == 1:  # q[4] = numero de estimacion
                            ws[acum_conceptos][8].value = q[0]  # AGREGA LA CANTIDAD EST
                            ws[acum_conceptos][9].value = q[3]  # AGREGA EL IMPORTE EST
                        else:
                            ws[acum_conceptos][7 + q[4] + (q[4] - 1)].value = q[0]  # AGREGA LA CANTIDAD EST
                            ws[acum_conceptos][8 + q[4] + (q[4] - 1)].value = q[3]  # AGREGA EL IMPORTE EST

                    # SE AGREGAN LAS COLUMNAS DE ESTIMACION #
                numero_est = 1
                acum_estx = 0
                for r in range(estimacion_count):
                    numero_est += 1
                    if numero_est == 1:
                        pass
                    else:
                        acum_estx += 1
                        ws[9][7 + numero_est + acum_estx].value = 'ESTIMACION ' + str(numero_est)
                        ws[10][7 + numero_est + acum_estx].value = 'CANTIDAD'
                        ws[10][8 + numero_est + acum_estx].value = 'IMPORTE'''


            # wb.save("/tmp/finiquito.xlsx")

            f = open('/tmp/finiquito.xlsx', mode="rb")
            return http.request.make_response(f.read(),
                                            [('Content-Type', 'application/octet-stream'),
                                            ('Content-Disposition',
                                                'attachment; filename="{}"'.format('finiquito.xlsx'))
                                            ])

        except Exception as e:
            return "Upss! algo salio mal en: " + str(e)


'''class FiniquitoExcel(http.Controller):
    @http.route('/finiquito_excel/finiquito_excel/', auth='public')
    def index(self, **kw):
        try:
            partidax = http.request.env['partidas.partidas']
            partida = partidax.sudo().search([('id', '=', kw['id'])])
            # workbook = load_workbook(filename="/home/gerardo/Developments/odoo12/extra-addons/finiquito_excel/static/plantilla.xlsx")
            workbook = load_workbook(
                filename="/usr/lib/python3/dist-packages/odoo/odoo-extra-addons/finiquito_excel/static/plantilla.xlsx")
            sheet = workbook.active
            wb = Workbook()
            nombre_partida = ''
            contratista = ''
            total_contrato = ''
            fill = PatternFill(fill_type=None, start_color='bdbdbd', end_color='bdbdbd')
            double = Side(border_style="double", color="000000")
            thin = Side(border_style="thin", color="000000")

            # ESTILOS DE SUBTOTAL
            subtotal_estilo = NamedStyle(name="subtotal_estilo")
            subtotal_estilo.font = Font(bold=True, size=9)
            subtotal_estilo.fill = PatternFill("solid", fgColor="ff7043")
            subtotal_estilo.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            wb.add_named_style(subtotal_estilo)

            # ESTILOS DE CATEGORIAS
            categoria_estilo = NamedStyle(name="categoria_estilo")
            categoria_estilo.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            categoria_estilo.font = Font(bold=True, size=9)
            categoria_estilo.fill = PatternFill("solid", fgColor="bdbdbd")
            wb.add_named_style(categoria_estilo)

            # ESTILOS DE HEADER DE ESTIMACIONES
            header_est_estilo = NamedStyle(name="header_est_estilo")
            header_est_estilo.font = Font(name='Arial', size=9, bold=True)
            header_est_estilo.fill = PatternFill("solid", fgColor="ffff00")
            header_est_estilo.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            wb.add_named_style(header_est_estilo)

            # ESTILOS TOTAL
            total_estilo = NamedStyle(name="total_estilo")
            total_estilo.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            total_estilo.fill = PatternFill("solid", fgColor="e65100")
            total_estilo.font = Font(color="fafafa", name='Arial', size=9, bold=True)
            wb.add_named_style(total_estilo)

            for i in partida:
                print(' INICIO ')
                b_partida = http.request.env['partidas.partidas'].browse(i.id)
                nombre_partida = b_partida.nombre_partida
                contratista = b_partida.contratista.name
                total_contrato = b_partida.total
                estimacion = http.request.env['control.estimaciones'].search(
                    [('obra.id', '=', b_partida.id)])  # , ('idobra', '=', '1')
                estimacion_c = http.request.env['control.estimaciones'].search_count(
                    [('obra.id', '=', b_partida.id)])  # , ('idobra', '=', '1')
                acum = 0
                aviso = 0
                posicionador = 0
                contador_est = 0
                # CICLO PARA LOS CONCEPTOS DEL CATALOGO
                for y in b_partida.conceptos_partidas:

                    if str(y.medida) == 'False':
                        y.medida = ''
                    if str(y.cantidad) == '0.0':
                        y.cantidad = ''
                    if str(y.precio_unitario) == '0.0':
                        y.precio_unitario = ''
                    if str(y.importe) == '0.0':
                        y.importe = ''

                    # AVISO INDICA CUANDO HAY QUE APLICAR SUBTOTAL
                    acum += 1 + aviso
                    column_letter = get_column_letter(2)  # LETRA CLAVE
                    column_letter_concepto = get_column_letter(3)  # LETRA CONCEPTO
                    posicionador = acum + 10  # VARIABLE ACUMULABLE PARA EL POSICIONAMIENTO DE LAS CELDAS
                    aviso = 0
                    print(y.clave_linea, ' ----------------------- ', y.precio_unitario)

                    sheet[column_letter_concepto + str(posicionador)] = str(y.concepto)
                    sheet[column_letter + str(posicionador)] = str(y.clave_linea)
                    sheet[column_letter + str(posicionador)].border = Border(top=thin, left=thin, right=thin,
                                                                             bottom=thin)
                    sheet[column_letter_concepto + str(posicionador)].border = Border(top=thin, left=thin, right=thin,
                                                                                      bottom=thin)

                    cantidad_estimado = 0
                    contador_est = 0
                    acum_estimado = 0

                    continuacion_cantidad = ''
                    continuacion_importe = ''

                    columna_total_estimados = ''
                    columna_total_cantidad = ''
                    for x in estimacion:
                        acum_estimado += 2
                        contador_est += 1
                        # VARIABLES
                        column_concepto_estimado = get_column_letter(6 + acum_estimado)  # CONCEPTO ESTIMACION
                        column_concepto_importe = get_column_letter(7 + acum_estimado)  # CONCEPTO ESTIMACION

                        columna_total_estimados = get_column_letter(9 + acum_estimado)  # TOTAL ESTIMADO
                        columna_total_cantidad = get_column_letter(8 + acum_estimado)  # TOTAL ESTIMADO

                        if not y.precio_unitario or str(y.precio_unitario) == '':  # SI ES CATEGORIA
                            # aviso = 1
                            sheet[column_letter + str(posicionador)] = str(y.clave_linea)
                            sheet[column_letter + str(posicionador)].alignment = Alignment(wrap_text=True)
                            sheet[column_letter + str(posicionador)].style = categoria_estilo
                            sheet[column_letter_concepto + str(posicionador)] = str(y.concepto)
                            sheet[column_letter_concepto + str(posicionador)].style = categoria_estilo

                        # HEADERS ESTIMACIONES VARIABLES
                        estimacion_header = get_column_letter(6 + acum_estimado)
                        column_letterest = get_column_letter(7 + acum_estimado)
                        estimacion_header_importe = get_column_letter(7 + acum_estimado)

                        b_est = http.request.env['control.estimaciones'].browse(x.id)
                        detalle_concepto = http.request.env['control.detalle_conceptos'].search(
                            [('id_partida.id', '=', b_partida.id), ('num_est', '=', b_est.idobra)
                                , ('clave_linea', '=', y.clave_linea)])

                        # BUSQUEDA DEL CONCEPTO E IMPRESION
                        for j in detalle_concepto:
                            # HEADER ESTIMACIONES
                            sheet[estimacion_header + str(10)] = 'CANTIDAD'  # ESTIMACION
                            sheet[estimacion_header + str(10)].style = header_est_estilo
                            sheet[estimacion_header + '9'] = 'ESTIMACION ' + str(b_est.idobra)
                            sheet[estimacion_header + '9'].alignment = Alignment(horizontal="center", vertical="center")
                            sheet.merge_cells("" + estimacion_header + "9:" + column_letterest + "9")
                            sheet[column_letterest + '9'].fill = PatternFill("solid", fgColor="ffff00")
                            sheet[estimacion_header + '9'].fill = PatternFill("solid", fgColor="ffff00")
                            sheet[estimacion_header + '9'].font = Font(name='Arial', size=9, bold=True)
                            sheet[estimacion_header + '10'].fill = PatternFill("solid", fgColor="ffff00")
                            sheet[estimacion_header + '9'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            sheet[estimacion_header_importe + str(10)] = 'IMPORTE'  # ESTIMACION
                            sheet[estimacion_header_importe + str(10)].style = header_est_estilo

                            detalle = http.request.env['control.detalle_conceptos'].browse(j.id)

                            # print(y.clave_linea, '-----xxxxxxxxxxxxxxxxxx--------')
                            # if y.precio_unitario == 0 or not y.precio_unitario or str(y.precio_unitario) == '0.0':
                            # else:
                            # print(detalle.clave_linea, y.clave_linea, detalle.num_est, ' p ', detalle.precio_unitario)
                            cantidad_estimado = detalle.estimacion  # CANTIDAD ESTIMACION
                            importe_ejecutado = detalle.importe_ejecutado  # IMPORTE EJECUTADO
                            # ESTIMADO
                            sheet[column_concepto_estimado + str(posicionador)] = float(cantidad_estimado)
                            sheet[column_concepto_importe + str(posicionador)] = float(importe_ejecutado)
                            sheet[column_concepto_importe + str(
                                posicionador)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'

                            # # TOTALES DE CADA CONCEPTO EN T.IMPORTE -------------------------------------------

                            continuacion_importe += str(column_concepto_importe + str(posicionador)) + ','
                            continuacion_cantidad += str(column_concepto_estimado + str(posicionador)) + ','

                            # continuacion += ',' + str(column_lettery2 + str(acum_clavex + 9))

                            # sheet[columna_total_estimados + str(posicionador)] = "=SUM(" + column_concepto_importe + str(posicionador) + ":" + column_letter_timporte + str(acum_clavex + 8) + ")"
                            sheet[columna_total_estimados + str(posicionador)] = "=SUM(" + continuacion_importe + ")"
                            sheet[columna_total_estimados + str(
                                posicionador)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                            sheet[columna_total_cantidad + str(posicionador)] = "=SUM(" + continuacion_cantidad + ")"
                            sheet[columna_total_cantidad + str(
                                posicionador)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                        sheet[column_concepto_estimado + str(posicionador)].border = Border(top=thin, left=thin,
                                                                                            right=thin, bottom=thin)
                        sheet[column_concepto_importe + str(posicionador)].border = Border(top=thin, left=thin,
                                                                                           right=thin, bottom=thin)
                    sheet[columna_total_estimados + str(posicionador)].style = subtotal_estilo
                    sheet[columna_total_cantidad + str(posicionador)].style = subtotal_estilo

                    # CLAVE NORMAL
                    # sheet[column_letter + str(posicionador)] = str(y.clave_linea)
                    # sheet[column_letter_concepto + str(posicionador)] = str(y.concepto)
                    # sheet[column_letter_concepto + str(posicionador)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    # sheet[column_letter + str(posicionador)].border = Border(top=thin, left=thin, right=thin, bottom=thin)

                    letra_totales = get_column_letter(8 + acum_estimado)
                    letra_totales2 = get_column_letter(9 + acum_estimado)
                    column_letterq = get_column_letter(9 + acum_estimado)
                    column_letterz = get_column_letter(8 + acum_estimado)
                    # escribe en la celda
                    sheet[letra_totales + str(10)] = 'T.CANTIDAD'
                    sheet[letra_totales + str(10)].style = total_estilo
                    sheet[letra_totales + '9'] = 'TOTALES'
                    sheet[letra_totales + '9'].alignment = Alignment(horizontal="center", vertical="center")
                    sheet.merge_cells("" + column_letterz + "9:" + column_letterq + "9")
                    sheet[column_letterq + '9'].fill = PatternFill("solid", fgColor="e65100")
                    sheet[letra_totales + '9'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    sheet[letra_totales + '9'].fill = PatternFill("solid", fgColor="e65100")
                    sheet[letra_totales + '9'].font = Font(color="fafafa", name='Arial', size=9, bold=True)

                    column_letter = get_column_letter(9 + acum_estimado)
                    sheet[letra_totales2 + str(10)] = 'T.IMPORTE'
                    sheet[letra_totales2 + str(10)].style = total_estilo

                    # UNIDAD
                    column_letter_unidad = get_column_letter(4)
                    # posicionador = acum + 10 + aviso

                    sheet[column_letter_unidad + str(posicionador)] = str(y.medida)
                    sheet[column_letter_unidad + str(posicionador)].border = Border(top=thin, left=thin, right=thin,
                                                                                    bottom=thin)
                    if not y.precio_unitario or str(y.precio_unitario) == '':
                        sheet[column_letter_unidad + str(posicionador)].style = categoria_estilo

                    # CANTIDAD
                    letra = 0

                    for column in range(5, 6):
                        column_letter = get_column_letter(column)
                        posicionador = acum + 10 + aviso
                        if not y.precio_unitario or str(y.precio_unitario) == '':
                            sheet[column_letter + str(posicionador)].style = categoria_estilo
                        else:
                            letra += 1
                            sheet[column_letter + str(posicionador)] = y.cantidad
                            sheet[column_letter + str(posicionador)].border = Border(top=thin, left=thin, right=thin,
                                                                                     bottom=thin)

                    # PRECIO UNITARIO
                    for column in range(6, 7):
                        column_letter = get_column_letter(column)
                        posicionador = acum + 10 + aviso
                        if not y.precio_unitario or str(y.precio_unitario) == '':
                            sheet[column_letter + str(posicionador)].style = categoria_estilo
                            
                        else:
                            letra += 1
                            sheet[column_letter + str(posicionador)] = y.precio_unitario
                            sheet[column_letter + str(posicionador)].border = Border(top=thin, left=thin, right=thin,
                                                                                     bottom=thin)

                    # IMPORTE
                    for column in range(7, 8):
                        column_letter = get_column_letter(column)
                        # escribe en la celda
                        posicionador = acum + 10 + aviso
                        if not y.precio_unitario or str(y.precio_unitario) == '':
                            sheet[column_letter + str(posicionador)].style = categoria_estilo
                           
                        else:
                            letra += 1
                            sheet[column_letter + str(posicionador)] = y.importe
                            sheet[column_letter + str(posicionador)].border = Border(top=thin, left=thin, right=thin,
                                                                                     bottom=thin)
                contador_est = 0

                # SUBTOTAL ULTIMO
                acum_estimado_total = 0
                for x in estimacion:
                    b_est = http.request.env['control.estimaciones'].browse(x.id)
                    acum_estimado_total += b_est.estimado

                column_letter = get_column_letter(2)
                column_letter_1 = get_column_letter(5)
                column_letter_2 = get_column_letter(6)
                column_letter_3 = get_column_letter(7)
                sheet[letra_totales + str(posicionador + 1)] = 'GRAN TOTAL'
                sheet[letra_totales2 + str(posicionador + 1)] = acum_estimado_total
                sheet[letra_totales2 + str(posicionador + 1)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                # sheet[column_letter + str(posicionador + 1)].fill = PatternFill("solid", fgColor="ff7043")
                sheet[letra_totales + str(posicionador + 1)].style = subtotal_estilo
                sheet[letra_totales2 + str(posicionador + 1)].style = subtotal_estilo
           
                # CANTIDAD, PRECIO U., IMPORTE
                # sheet[column_letter_3 + str(posicionador + 1)] = "=SUM(G" + str(variable3) + ":G" + str(posicionador) + ")"

            for column in range(7, 8):
                column_letter = get_column_letter(column)
                # escribe en la celda
                sheet[column_letter + "7"] = nombre_partida

            for column in range(2, 3):
                column_letter = get_column_letter(column)
                # escribe en la celda
                sheet[column_letter + "7"] = contratista

            # Save the spreadsheet
            workbook.save("/tmp/finiquito.xlsx")

            # prs.save('/tmp/test.pptx')
            f = open('/tmp/finiquito.xlsx', mode="rb")
            return http.request.make_response(f.read(),
                                            [('Content-Type', 'application/octet-stream'),
                                            ('Content-Disposition',
                                                'attachment; filename="{}"'.format('finiquito.xlsx'))
                                            ])
        
        except Exception as e:
            return "Upss! algo salio mal en: " + str(e)'''


class ConceptosEstimados(http.Controller):
    @http.route('/finiquito_excel_conceptos/finiquito_excel_conceptos/', auth='public')
    def index(self, **kw):
        try:
            estimacion = http.request.env['control.estimaciones'].search([('id', '=', kw['id'])])  # , ('idobra', '=', '1')
            # estimacion_c = http.request.env['control.estimaciones'].search_count([('obra.id', '=', kw['id'])])  # , ('idobra', '=', '1')
            # workbook = load_workbook(filename="/home/gerardo/Developments/odoo12/extra-addons/finiquito_excel/static/conceptos_estimados_plantilla.xlsx")
            workbook = load_workbook(filename="/usr/lib/python3/dist-packages/odoo/odoo-extra-addons/finiquito_excel/static/conceptos_estimados_plantilla.xlsx")
            sheet = workbook.active
            wb = Workbook()

            fill = PatternFill(fill_type=None, start_color='bdbdbd', end_color='bdbdbd')
            double = Side(border_style="double", color="000000")
            thin = Side(border_style="thin", color="000000")

            nombre_partida = ''
            contratista = ''
            idobra = ''
            periodo_del = ''
            periodo_al = ''
            residente = ''

            for o in estimacion:
                estimacionx = http.request.env['control.estimaciones'].browse(o.id)

                partidax = http.request.env['partidas.partidas']
                partida = partidax.sudo().search([('id', '=', estimacionx.obra.id)])

                for x in partida:
                    b_partida = http.request.env['partidas.partidas'].browse(x.id)
                    for p in b_partida.residente_obra:
                        residente = p.name
                    nombre_partida = b_partida.nombre_partida
                    contratista = b_partida.contratista.name

                acum = 0
                pos_concepto = 0
                aviso = 0
                xd = 0
                for i in estimacionx:
                    idobra = i.idobra
                    periodo_del = i.fecha_inicio_estimacion
                    periodo_al = i.fecha_termino_estimacion
                    for y in estimacionx.conceptos_partidas:
                        pos_concepto += 1
                        if str(y.medida) == 'False':
                            y.medida = ''
                        if str(y.cantidad) == '0.0':
                            y.cantidad = ''
                        if str(y.precio_unitario) == '0.0':
                            y.precio_unitario = ''
                        if str(y.importe_ejecutado) == '0.0':
                            y.importe_ejecutado = ''
                        # AVISO INDICA CUANDO HAY QUE APLICAR SUBTOTAL
                        # CLAVE
                        if y.estimacion > 0 or y.estimacion < 0:
                            acum += 1 + aviso
                            # CLAVE
                            for column in range(2, 3):
                                column_letter = get_column_letter(column)
                                xd = acum + 11
                                aviso = 0
                                sheet[column_letter + str(xd)] = str(y.clave_linea)
                                sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            # CONCEPTO
                            for column in range(3, 4):
                                column_letter = get_column_letter(column)
                                xd = acum + 11
                                aviso = 0
                                sheet[column_letter + str(xd)] = str(y.concepto)
                                sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            # MEDIDA
                            for column in range(4, 5):
                                column_letter = get_column_letter(column)
                                xd = acum + 11
                                aviso = 0
                                sheet[column_letter + str(xd)] = str(y.medida)
                                sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            # ESTIMADO
                            for column in range(5, 6):
                                column_letter = get_column_letter(column)
                                xd = acum + 11
                                aviso = 0
                                sheet[column_letter + str(xd)] = str(y.estimacion)
                                sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            # PRECIO UNITARIO
                            for column in range(6, 7):
                                column_letter = get_column_letter(column)
                                xd = acum + 11
                                aviso = 0
                                sheet[column_letter + str(xd)] = y.precio_unitario
                                sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                                sheet[column_letter + str(xd)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                            # IMPORTE
                            for column in range(7, 8):
                                column_letter = get_column_letter(column)
                                xd = acum + 11
                                aviso = 0
                                sheet[column_letter + str(xd)] = y.importe_ejecutado
                                sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                                sheet[column_letter + str(xd)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                        else:
                            pass

                for column in range(3, 4):
                    column_letter = get_column_letter(column)
                    sheet[column_letter + str(xd + 4)] = residente

                column_letteru = get_column_letter(6)  # TOTAL
                column_letterx = get_column_letter(7)  # TOTAL
                sheet[column_letteru + str(xd + 1)] = 'TOTAL'
                sheet[column_letteru + str(xd + 1)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                sheet[column_letteru + str(xd + 1)].fill = PatternFill("solid", fgColor="ff7043")
                sheet[column_letteru + str(xd + 1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                sheet[column_letterx + str(xd + 1)] = "=SUM(" + column_letterx + str(12) + ":" + column_letterx + str(xd) + ")"
                sheet[column_letterx + str(xd + 1)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                sheet[column_letterx + str(xd + 1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            for column in range(6, 7):
                column_letter = get_column_letter(column)
                sheet[column_letter + "8"] = nombre_partida

            for column in range(3, 4):
                column_letter = get_column_letter(column)
                sheet[column_letter + "8"] = contratista

            for column in range(7, 8):
                column_letter = get_column_letter(column)
                sheet[column_letter + "6"] = idobra

            for column in range(4, 5):
                column_letter = get_column_letter(column)
                sheet[column_letter + "8"] = periodo_del
                sheet[column_letter + "9"] = periodo_al
            
            # Save the spreadsheet
            workbook.save("/tmp/conceptos_estimacion.xlsx")

            # prs.save('/tmp/test.pptx')
            f = open('/tmp/conceptos_estimacion.xlsx', mode="rb")
            return http.request.make_response(f.read(),
                                            [('Content-Type', 'application/octet-stream'),
                                            ('Content-Disposition',
                                                'attachment; filename="{}"'.format('conceptos_estimacion.xlsx'))
                                            ])
        
        except Exception as e:
            return "Upss! algo salio mal en: " + str(e)

        