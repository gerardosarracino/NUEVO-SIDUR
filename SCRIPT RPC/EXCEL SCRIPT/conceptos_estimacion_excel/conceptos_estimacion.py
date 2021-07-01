from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import odoorpc, csv

from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

usuario = 'admin'
password = 'spiderboras'
odoo = odoorpc.ODOO('sidur.galartec.com', port=8069)
odoo.login('sidur2020', usuario, password)

partida = odoo.env['partidas.partidas'].search([('nombre_partida', '=', 'SIDUR-ED-20-003.1871')])

nombre_partida = ''
contratista = ''
idobra = ''
periodo_del = ''
periodo_al = ''
residente = ''

workbook = load_workbook(filename="ok.xlsx")
sheet = workbook.active
wb = Workbook()

fill = PatternFill(fill_type=None,  start_color='bdbdbd', end_color='bdbdbd')
double = Side(border_style="double", color="000000")
thin = Side(border_style="thin", color="000000")

for i in partida:
    b_partida = odoo.env['partidas.partidas'].browse(i)

    for p in b_partida.residente_obra:
        residente = p.name

    nombre_partida = b_partida.nombre_partida
    contratista = b_partida.contratista.name

    estimacion = odoo.env['control.estimaciones'].search([('obra.id', '=', b_partida.id), ('idobra', '=', '1')]) # , ('idobra', '=', '1')
    estimacion_c = odoo.env['control.estimaciones'].search_count([('obra.id', '=', b_partida.id), ('idobra', '=', '1')]) # , ('idobra', '=', '1')
    acum = 0
    pos_concepto = 0
    aviso = 0
    variable1 = ''
    variable2 = ''
    variable3 = 0
    variable4 = 0
    xd = 0

    for i in estimacion:
        estimacionx = odoo.env['control.estimaciones'].browse(i)
        idobra = estimacionx.idobra
        periodo_del = estimacionx.fecha_inicio_estimacion
        periodo_al = estimacionx.fecha_termino_estimacion
        for y in estimacionx.conceptos_partidas:
            pos_concepto += 1
            if str(y.medida) == 'False':
                y.medida = ''
            if str(y.cantidad) == '0.0':
                y.cantidad = ''
            if str(y.precio_unitario) == '0.0':
                y.precio_unitario = ''
            if str(y.importe) == '0.0':
                y.importe = ''
            # AVISO INDICA CUANDO HAY QUE APLICAR SUBTOTAL
            # CLAVE
            if y.estimacion > 0:
                acum += 1 + aviso
                # CLAVE
                for column in range(2, 3):
                    column_letter = get_column_letter(column)
                    xd = acum + 11
                    aviso = 0
                    sheet[column_letter + str(xd)] = str(y.clave_linea)
                    sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                # CONCEPTO
                for column in range(3, 4):
                    column_letter = get_column_letter(column)
                    xd = acum + 11
                    aviso = 0
                    sheet[column_letter + str(xd)] = str(y.concepto)
                    sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                # MEDIDA
                for column in range(4, 5):
                    column_letter = get_column_letter(column)
                    xd = acum + 11
                    aviso = 0
                    sheet[column_letter + str(xd)] = str(y.medida)
                    sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                # ESTIMADO
                for column in range(5, 6):
                    column_letter = get_column_letter(column)
                    xd = acum + 11
                    aviso = 0
                    sheet[column_letter + str(xd)] = str(y.estimacion)
                    sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                # PRECIO UNITARIO
                for column in range(6, 7):
                    column_letter = get_column_letter(column)
                    xd = acum + 11
                    aviso = 0
                    sheet[column_letter + str(xd)] = y.precio_unitario
                    sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    sheet[column_letter + str(xd)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                # IMPORTE
                for column in range(7, 8):
                    column_letter = get_column_letter(column)
                    xd = acum + 11
                    aviso = 0
                    sheet[column_letter + str(xd)] = y.importe_ejecutado
                    sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    sheet[column_letter + str(xd)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
            else:
                pass

    for column in range(3,4):
        column_letter = get_column_letter(column)
        sheet[column_letter + str(xd+4)] = residente

    column_letteru = get_column_letter(6)  # TOTAL
    column_letterx = get_column_letter(7)  # TOTAL
    sheet[column_letteru + str(xd + 1)] = 'TOTAL'
    sheet[column_letteru + str(xd + 1)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
    sheet[column_letteru + str(xd + 1)].fill = PatternFill("solid", fgColor="ff7043")
    sheet[column_letteru + str(xd + 1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    sheet[column_letterx + str(xd + 1)] = "=SUM(" + column_letterx + str(12) + ":" + column_letterx + str(xd) + ")"
    sheet[column_letterx + str(xd + 1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
print('exito')
for column in range(6,7):
    column_letter = get_column_letter(column)
    sheet[column_letter + "8"] = nombre_partida

for column in range(3,4):
    column_letter = get_column_letter(column)
    sheet[column_letter + "8"] = contratista

for column in range(7,8):
    column_letter = get_column_letter(column)
    sheet[column_letter + "6"] = idobra

for column in range(4,5):
    column_letter = get_column_letter(column)
    sheet[column_letter + "8"] = periodo_del
    sheet[column_letter + "9"] = periodo_al

# Save the spreadsheet
workbook.save(filename="resultado.xlsx")
