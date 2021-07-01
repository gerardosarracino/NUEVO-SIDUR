from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import odoorpc, csv

from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

usuario = 'admin'
password = 'Sp1d3rb0r4s2020'
odoo = odoorpc.ODOO('sidur.galartec.com', port=8069)
odoo.login('sidur2020', usuario, password)

partida = odoo.env['partidas.partidas'].search([('nombre_partida', '=', 'SIDUR-ED-19-133.1779')])

nombre_partida = ''
contratista = ''
total_contrato = ''

workbook = load_workbook(filename="ok.xlsx")
sheet = workbook.active
wb = Workbook()


fill = PatternFill(fill_type=None,  start_color='bdbdbd', end_color='bdbdbd')
double = Side(border_style="double", color="000000")
thin = Side(border_style="thin", color="000000")

for i in partida:
    print(' INICIO ')
    b_partida = odoo.env['partidas.partidas'].browse(i)
    nombre_partida = b_partida.nombre_partida
    contratista = b_partida.contratista.name
    total_contrato = b_partida.total

    estimacion = odoo.env['control.estimaciones'].search([('obra.id', '=', b_partida.id), ('idobra', '<', '3')]) # , ('idobra', '=', '1')
    estimacion_c = odoo.env['control.estimaciones'].search_count([('obra.id', '=', b_partida.id), ('idobra', '<', '3')]) # , ('idobra', '=', '1')
    acum = 0
    pos_concepto = 0
    aviso = 0
    variable1 = ''
    variable2 = ''
    variable3 = 0
    variable4 = 0
    variable5 = 0
    variable_est = 0
    xd = 0

    acum_clave = 0
    acum_clave_ = 0
    acum_clavexxx = 0  # CATEGORIA

    clave_sub = 0
    contador_est = 0

    # T.CANTIDAD T.TOTAL
    pos_total1 = 8 + (estimacion_c * 2)
    pos_total2 = 8 + (estimacion_c * 2) + 1

    letra_subtotal_totales = get_column_letter(pos_total1)
    column_letter_timporte = get_column_letter(pos_total2)
    clave_inicial = get_column_letter(2)
    acum_subtotal_totales = 0
    clave_subx = 0
    clave = ''
    acum_linea_total = 0

    for y in b_partida.conceptos_partidas:
        acum_subtotal_totales += 1 + clave_subx

        print(y.clave_linea, ' ----------------------- ', y.precio_unitario)
        acum_clave += 1 + clave_sub
        acum_clavexxx += 1

        if contador_est == 1 or contador_est == 0:
            estx = 8
            esty = 9
            imporx = 9
            impory = 10
            va = 10
        else:
            estx = int(va)
            esty = int(estx + 1)
            imporx = int(esty)
            impory = int(esty + 1)

        pos_concepto += 1
        if str(y.medida) == 'False':
            y.medida = ''
        if str(y.cantidad) == '0.0':
            y.cantidad = ''
        if str(y.precio_unitario) == '0.0':
            y.precio_unitario = ''
        if str(y.importe) == '0.0':
            y.importe = ''

        # AVISO INDICA CUANDO HAY QUE APLICAR SUBTOTAL
        acum += 1 + aviso
        # CLAVE
        for column in range(2, 3):
            column_letter = get_column_letter(column) # LETRA CLAVE
            column_letter_concepto = get_column_letter(3) # LETRA CONCEPTO

            xd = acum + 10

            aviso = 0

            # sheet[column_letter_concepto + str(xd)] = str(y.concepto)
            # sheet[column_letter_concepto + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if not y.precio_unitario or str(y.precio_unitario) == '':
                if pos_concepto > 2:


                    # SUBTOTALES
                    sheet[column_letter + str(xd)] = 'SUBTOTAL'
                    sheet[column_letter + str(xd)].fill = PatternFill("solid", fgColor="ff7043")
                    sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    aviso = 1
                    # SIGUIENTE CATEGORIA CALVE
                    sheet[column_letter + str(xd + 1)] = str(y.clave_linea)
                    sheet[column_letter + str(xd + 1)].alignment = Alignment(wrap_text=True)
                    sheet[column_letter + str(xd + 1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    sheet[column_letter + str(xd + 1)].font = Font(bold=True, size=9)
                    sheet[column_letter + str(xd + 1)].fill = PatternFill("solid", fgColor="bdbdbd")

                    # CONCEPTO
                    sheet[column_letter_concepto + str(xd + 1)] = str(y.concepto)
                    sheet[column_letter_concepto + str(xd+1)].font = Font(bold=True, size=9)
                    sheet[column_letter_concepto + str(xd+1)].fill = PatternFill("solid", fgColor="bdbdbd")
                    sheet[column_letter_concepto + str(xd+1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    sheet[column_letter_concepto + str(xd)].fill = PatternFill("solid", fgColor="ff7043")
                    sheet[column_letter_concepto + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)

                    # ESTIMADOS
                    acum_est = 0
                    for h in range(estimacion_c):
                        acum_est += 2
                        column_letter_estix = get_column_letter(6 + acum_est)  # COLUMNAS DIBUJAR ESTIMADOS
                        column_letter_estiy = get_column_letter(7 + acum_est)  # COLUMNAS DIBUJAR ESTIMADOS

                        letra_importe = column_letter_estiy
                        # SUB

                        sheet[column_letter_estix + str(xd+1)].fill = PatternFill("solid", fgColor="bdbdbd")
                        sheet[column_letter_estiy + str(xd+1)].fill = PatternFill("solid", fgColor="bdbdbd")
                        sheet[column_letter_estix + str(xd)].fill = PatternFill("solid", fgColor="ff7043")
                        sheet[column_letter_estiy + str(xd)].fill = PatternFill("solid", fgColor="ff7043")
                        sheet[column_letter_estiy + str(xd)] = "=SUM(" + letra_importe + str(variable_est) + ":" + letra_importe + str(xd - 1) + ")"
                        sheet[column_letter_estiy + str(xd)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                        sheet[column_letter_estix + str(xd+1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        sheet[column_letter_estiy + str(xd+1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        sheet[column_letter_estix + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        sheet[column_letter_estiy + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)

                    variable_est = xd + 2
                else:

                    variable_est = xd

                    # CLAVES
                    sheet[column_letter + str(xd)] = str(y.clave_linea)
                    sheet[column_letter + str(xd)].alignment = Alignment(wrap_text=True)
                    sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    sheet[column_letter + str(xd)].font = Font(bold=True, size=9)
                    sheet[column_letter + str(xd)].fill = PatternFill("solid", fgColor="bdbdbd")

                    sheet[column_letter_concepto + str(xd)] = str(y.concepto)
                    sheet[column_letter_concepto + str(xd)].font = Font(bold=True, size=9)
                    sheet[column_letter_concepto + str(xd)].fill = PatternFill("solid", fgColor="bdbdbd")
                    sheet[column_letter_concepto + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    # sheet[column_letter_concepto + str(xd)].fill = PatternFill("solid", fgColor="ff7043")

                    # ESTIMADOS
                    acum_est = 0
                    for h in range(estimacion_c):
                        acum_est += 2
                        column_letter_estix = get_column_letter(6 + acum_est)  # COLUMNAS DIBUJAR ESTIMADOS
                        column_letter_estiy = get_column_letter(7 + acum_est)  # COLUMNAS DIBUJAR ESTIMADOS
                        sheet[column_letter_estix + str(xd)].fill = PatternFill("solid", fgColor="bdbdbd")
                        sheet[column_letter_estiy + str(xd)].fill = PatternFill("solid", fgColor="bdbdbd")
                        sheet[column_letter_estix + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        sheet[column_letter_estiy + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            else:
                cantidad_estimado = 0
                contador_est = 0
                acum_estimado = 0

                clave_inicial = get_column_letter(2)
                clave = sheet[clave_inicial + str(acum_clave + 9)].value

                for x in estimacion:
                    acum_estimado += 2
                    contador_est += 1
                    # VARIABLES
                    column_concepto_estimado = get_column_letter(6+acum_estimado)  # CONCEPTO ESTIMACION
                    column_concepto_importe = get_column_letter(7+acum_estimado)  # CONCEPTO ESTIMACION

                    # HEADERS ESTIMACIONES VARIABLES
                    estimacion_header = get_column_letter(6+acum_estimado)
                    column_letterest = get_column_letter(7+acum_estimado)
                    estimacion_header_importe = get_column_letter(7+acum_estimado)

                    b_est = odoo.env['control.estimaciones'].browse(x)
                    detalle_concepto = odoo.env['control.detalle_conceptos'].search(
                        [('id_partida.id', '=', b_partida.id), ('num_est', '=', b_est.idobra)
                         , ('clave_linea', '=', y.clave_linea)])

                    # SI LA LINEA ES SUBTOTAL IGNORAR
                    '''if str(clave) == 'SUBTOTAL':
                        print('SUBOTTAL')
                        clave_sub = 1
                        print('SUBOTTAL')
                        column_letter_esti = get_column_letter(6 + acum_estimado)
                        print(column_letter_esti)
                        column_letter_esti2 = get_column_letter(7 + acum_estimado)

                        sheet[column_letter_esti + str(acum_clave + 10)].fill = PatternFill("solid", fgColor="bdbdbd")
                        sheet[column_letter_esti2 + str(acum_clave + 10)].fill = PatternFill("solid", fgColor="bdbdbd")
                        sheet[column_letter_esti + str(acum_clave + 10)].border = Border(top=thin, left=thin, right=thin,
                                                                            bottom=thin)
                        sheet[column_letter_esti2 + str(acum_clave + 10)].border = Border(top=thin, left=thin, right=thin,
                                                                            bottom=thin)
                        # -----------------  SUBTOTAL DE CANTIDAD
                        sheet[column_letter_esti2 + str(acum_clave + 9)].fill = PatternFill("solid", fgColor="ff7043")
                        sheet[column_letter_esti2 + str(acum_clave + 9)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        sheet[column_letter_esti + str(acum_clave + 9)].fill = PatternFill("solid", fgColor="ff7043")
                        sheet[column_letter_esti + str(acum_clave + 9)].border = Border(top=thin, left=thin, right=thin, bottom=thin)'''

                    # BUSQUEDA DEL CONCEPTO E IMPRESION
                    for j in detalle_concepto:
                        # HEADER ESTIMACIONES
                        sheet[estimacion_header + str(10)] = 'CANTIDAD'
                        sheet[estimacion_header + str(10)].font = Font(name='Arial', size=9, bold=True)
                        sheet[estimacion_header + str(10)].fill = PatternFill("solid", fgColor="ffff00")
                        sheet[estimacion_header + '9'] = 'ESTIMACION ' + str(b_est.idobra)
                        sheet[estimacion_header + '9'].alignment = Alignment(horizontal="center", vertical="center")
                        sheet.merge_cells("" + estimacion_header + "9:" + column_letterest + "9")
                        sheet[column_letterest + '9'].fill = PatternFill("solid", fgColor="ffff00")
                        sheet[estimacion_header + '9'].fill = PatternFill("solid", fgColor="ffff00")
                        sheet[estimacion_header + '9'].font = Font(name='Arial', size=9, bold=True)
                        sheet[estimacion_header + '10'].fill = PatternFill("solid", fgColor="ffff00")
                        sheet[estimacion_header + '9'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        sheet[estimacion_header_importe + str(10)] = 'IMPORTE'
                        sheet[estimacion_header_importe + str(10)].font = Font(name='Arial', size=9, bold=True)
                        sheet[estimacion_header_importe + str(10)].fill = PatternFill("solid", fgColor="ffff00")
                        sheet[estimacion_header_importe + str(10)].border = Border(top=thin, left=thin, right=thin,
                                                                                   bottom=thin)

                        detalle = odoo.env['control.detalle_conceptos'].browse(j)

                        if y.precio_unitario == 0 or not y.precio_unitario or str(y.precio_unitario) == '0.0':
                            pass
                        else:
                            clave_sub = 0
                            xy = impory
                            print(detalle.clave_linea, y.clave_linea, detalle.num_est, ' p ', detalle.precio_unitario)
                            cantidad_estimado = detalle.estimacion # CANTIDAD ESTIMACION
                            importe_ejecutado = detalle.importe_ejecutado # IMPORTE EJECUTADO
                            # ESTIMADO
                            sheet[column_concepto_estimado + str(xd)] = float(cantidad_estimado)
                            sheet[column_concepto_importe + str(xd)] = float(importe_ejecutado)
                            sheet[column_concepto_importe + str(xd)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'

                # CLAVE NORMAL
                sheet[column_letter + str(xd)] = str(y.clave_linea)
                sheet[column_letter_concepto + str(xd)] = str(y.concepto)

                sheet[column_letter_concepto + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)

                # T.CANTIDAD T.TOTAL
                # categoria = sheet[clave_cat + str(acum_subtotal_totales + 10)].value
                clave = sheet[clave_inicial + str(acum_subtotal_totales + 9)].value
                if str(clave) == 'SUBTOTAL':
                    acum_linea_total += 1
                    clave_subx = 1
                    sheet[letra_subtotal_totales + str(acum_subtotal_totales + 9)].border = Border(top=thin, left=thin, right=thin,
                                                                                bottom=thin)
                    sheet[letra_subtotal_totales + str(acum_subtotal_totales + 9)].fill = PatternFill("solid", fgColor="ff7043")
                    sheet[column_letter_timporte + str(acum_subtotal_totales + 9)].border = Border(top=thin, left=thin,
                                                                                         right=thin, bottom=thin)
                    sheet[column_letter_timporte + str(acum_subtotal_totales + 9)].fill = PatternFill("solid",
                                                                                            fgColor="ff7043")
                    sheet[letra_subtotal_totales + str(acum_subtotal_totales + 10)].border = Border(top=thin, left=thin, right=thin,
                                                                                 bottom=thin)
                    sheet[letra_subtotal_totales + str(acum_subtotal_totales + 10)].fill = PatternFill("solid", fgColor="bdbdbd")
                    sheet[column_letter_timporte + str(acum_subtotal_totales + 10)].border = Border(top=thin, left=thin,
                                                                                          right=thin, bottom=thin)
                    sheet[column_letter_timporte + str(acum_subtotal_totales + 10)].fill = PatternFill("solid",
                                                                                             fgColor="bdbdbd")
                else:
                    continuacion = ''
                    continuacion_tim = ''
                    letra = 0
                    pasar = 0
                    pos = 8
                    column_letterx = ''
                    column_lettery = ''
                    aviso = 0
                    pos_tim = 9
                    column_letterx_tim = ''
                    column_lettery_tim = ''
                    # CICLO PARA ACUMULAR FORMULAS DE CADA LINEA
                    for i in range(estimacion_c):
                        pasar += 1
                        # pasar los ultimos dos registros
                        if pasar <= (estimacion_c - 2):
                            letra += 2
                            column_lettery2 = get_column_letter((pos + 2) + letra)
                            column_lettery2_tim = get_column_letter((pos_tim + 2) + letra)
                            # STRING ACUMULADO
                            continuacion += ',' + str(column_lettery2 + str(acum_subtotal_totales + 9))
                            continuacion_tim += ',' + str(column_lettery2_tim + str(acum_subtotal_totales + 9))
                        else:
                            pass
                    column_letterx = get_column_letter(pos)
                    column_lettery = get_column_letter(pos + 2)
                    column_letterx_tim = get_column_letter(pos_tim)
                    column_lettery_tim = get_column_letter(pos_tim + 2)

                letra_totales = get_column_letter(8+acum_estimado)
                column_letterq = get_column_letter(9+acum_estimado)
                # escribe en la celda
                sheet[letra_totales + str(10)] = 'T.CANTIDAD'
                sheet[letra_totales + str(10)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                sheet[letra_totales + str(10)].fill = PatternFill("solid", fgColor="e65100")
                sheet[letra_totales + str(10)].font = Font(color="fafafa", name='Arial', size=9, bold=True)
                sheet[letra_totales + '9'] = 'TOTALES '
                sheet[letra_totales + '9'].alignment = Alignment(horizontal="center", vertical="center")
                sheet.merge_cells("" + column_letter + "9:" + column_letterq + "9")
                sheet[column_letterq + '9'].fill = PatternFill("solid", fgColor="e65100")
                sheet[letra_totales + '9'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                sheet[letra_totales + '9'].fill = PatternFill("solid", fgColor="e65100")
                sheet[letra_totales + '9'].font = Font(color="fafafa", name='Arial', size=9, bold=True)

                column_letter = get_column_letter(9+acum_estimado)
                sheet[letra_totales + str(10)] = 'T.IMPORTE'
                sheet[letra_totales + str(10)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                sheet[letra_totales + str(10)].fill = PatternFill("solid", fgColor="e65100")
                sheet[letra_totales + str(10)].font = Font(color="fafafa", name='Arial', size=9, bold=True)

        # UNIDAD
        for column in range(4, 5):
            column_letter = get_column_letter(column)
            xd = acum + 10 + aviso
            aviso = 0
            sheet[column_letter + str(xd)] = str(y.medida)
            sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if not y.precio_unitario or str(y.precio_unitario) == '':
                if pos_concepto > 2:
                    sheet[column_letter + str(xd - 1)].fill = PatternFill("solid", fgColor="ff7043")
                    aviso = 1
                sheet[column_letter + str(xd)].fill = PatternFill("solid", fgColor="bdbdbd")
        # CANTIDAD
        letra = 0

        for column in range(5, 6):
            column_letter = get_column_letter(column)
            xd = acum + 10 + aviso
            aviso = 0
            if not y.precio_unitario or str(y.precio_unitario) == '':
                if pos_concepto > 2:
                    # sheet[column_letter + str(xd - 1)] = "=SUM(E" + str(variable1) + ":E" + str(xd - 2) + ")"
                    sheet[column_letter + str(xd - 1)].fill = PatternFill("solid", fgColor="ff7043")
                    sheet[column_letter + str(xd - 1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    sheet[column_letter + str(xd)].fill = PatternFill("solid", fgColor="bdbdbd")
                    aviso = 1
                    variable1 = str(xd+1) # ESTA VARIABLE ALMACENA LA POSICION INICIAL PARA LA SUMA
                else:
                    sheet[column_letter + str(xd)].fill = PatternFill("solid", fgColor="bdbdbd")
                    sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    variable1 = str(xd+1) # ESTA VARIABLE ALMACENA LA POSICION INICIAL PARA LA SUMA
            else:
                letra += 1
                sheet[column_letter + str(xd)] = y.cantidad
                sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # PRECIO UNITARIO
        for column in range(6, 7):
            column_letter = get_column_letter(column)
            xd = acum + 10 + aviso
            aviso = 0
            if not y.precio_unitario or str(y.precio_unitario) == '':
                if pos_concepto > 2:
                    # sheet[column_letter + str(xd - 1)] = "=SUM(F" + str(variable2) + ":F" + str(xd - 2) + ")"
                    sheet[column_letter + str(xd - 1)].fill = PatternFill("solid", fgColor="ff7043")
                    sheet[column_letter + str(xd - 1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    sheet[column_letter + str(xd)].fill = PatternFill("solid", fgColor="bdbdbd")
                    sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    aviso = 1
                    # ESTA VARIABLE ALMACENA LA POSICION INICIAL PARA LA SUMA
                    variable2 = str(xd + 1)
                else:
                    sheet[column_letter + str(xd)].fill = PatternFill("solid", fgColor="bdbdbd")
                    sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    # ESTA VARIABLE ALMACENA LA POSICION INICIAL PARA LA SUMA
                    variable2 = str(xd + 1)
            else:
                letra += 1
                sheet[column_letter + str(xd)] = y.precio_unitario
                sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # IMPORTE
        for column in range(7, 8):
            column_letter = get_column_letter(column)
            # escribe en la celda
            xd = acum + 10 + aviso
            aviso = 0
            if not y.precio_unitario or str(y.precio_unitario) == '':
                if pos_concepto > 2:
                    sheet[column_letter + str(xd - 1)] = "=SUM(G" + str(variable3) + ":G" + str(xd - 2) + ")"
                    sheet[column_letter + str(xd - 1)].fill = PatternFill("solid", fgColor="ff7043")
                    sheet[column_letter + str(xd - 1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    sheet[column_letter + str(xd)].fill = PatternFill("solid", fgColor="bdbdbd")
                    sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    aviso = 1
                    # ESTA VARIABLE ALMACENA LA POSICION INICIAL PARA LA SUMA
                    variable3 = str(xd + 1)
                else:
                    sheet[column_letter + str(xd)].fill = PatternFill("solid", fgColor="bdbdbd")
                    sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    # ESTA VARIABLE ALMACENA LA POSICION INICIAL PARA LA SUMA
                    variable3 = str(xd + 1)
            else:
                letra += 1
                sheet[column_letter + str(xd)] = y.importe
                sheet[column_letter + str(xd)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    contador_est = 0
    xy = 0

    # SUBTOTAL ULTIMO
    column_letter = get_column_letter(2)
    column_letter_1 = get_column_letter(5)
    column_letter_2 = get_column_letter(6)
    column_letter_3 = get_column_letter(7)
    sheet[column_letter + str(xd + 1)] = 'SUBTOTAL'
    sheet[column_letter + str(xd + 1)].fill = PatternFill("solid", fgColor="ff7043")
    sheet[column_letter + str(xd + 1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for column in range(3, 8):
        column_letter = get_column_letter(column)
        sheet[column_letter + str(xd + 1)] = ''
        sheet[column_letter + str(xd + 1)].fill = PatternFill("solid", fgColor="ff7043")
        sheet[column_letter + str(xd + 1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    # CANTIDAD, PRECIO U., IMPORTE
    sheet[column_letter_3 + str(xd + 1)] = "=SUM(G" + str(variable3) + ":G" + str(xd) + ")"

    # ----------------X-------------  ESTIMACIONES -------------X-------------
    '''xd2 = 0
    acum_est = 0
    xy = 0
    posicion_estimacion = 0
    acum_estimado = 0
    clave_cat = get_column_letter(6)
    for x in estimacion:
        contador_est += 1

        # CONDICION DE ACUMULADOR PARA DETECTAR LAS CASILLAS DE CANTIDAD ESTIMADA PARA LA SUMATORIA
        if contador_est == 1:
            pass
        else:
            posicion_estimacion += 2

        b_est = odoo.env['control.estimaciones'].browse(x)

        for b_estt in b_est:
            acum_estimado += b_estt.estimado

        if contador_est == 1:
            estx = 8
            esty = 9
            imporx = 9
            impory = 10
        else:
            estx = int(xy)
            esty = int(estx + 1)
            imporx = int(esty)
            impory = int(esty + 1)

        print(' LA ESTIMACION ES', b_est.idobra, )
        acumx = 0

        pos_concepto2 = 0
        kk = 0
        acumy = 0
        pos_concepto3 = 0
        for y in b_est.conceptos_partidas:  # CICLO PRINCIPAL DE CONCEPTOS DE ESTIMACION

            concepto = y.concepto
            # CONVERTIR 0'S A VACIOS
            if str(y.estimacion) == '0.0':
                y.estimacion = ''
            if str(y.importe_ejecutado) == '0.0':
                y.importe_ejecutado = ''

            acumx += 1
            # -----------------------  COLUMNAS DE LAS ESTIMACIONES -----------------------------
            for column in range(estx, esty):
                column_letter = get_column_letter(column)
                column_letterest = get_column_letter(column+1)
                # escribe en la celda
                num_pos = 10
                sheet[column_letter + str(num_pos)] = 'CANTIDAD'
                sheet[column_letter + str(num_pos)].font = Font(name='Arial', size=9, bold=True)
                sheet[column_letter + str(num_pos)].fill = PatternFill("solid", fgColor="ffff00")
                sheet[column_letter + '9'] = 'ESTIMACION ' + str(b_est.idobra)
                sheet[column_letter + '9'].alignment = Alignment(horizontal="center", vertical="center")
                sheet.merge_cells("" + column_letter + "9:" + column_letterest + "9")
                sheet[column_letterest + '9'].fill = PatternFill("solid", fgColor="ffff00")
                sheet[column_letter + '9'].fill = PatternFill("solid", fgColor="ffff00")
                sheet[column_letter + '9'].font = Font(name='Arial', size=9, bold=True)
                sheet[column_letter + '10'].fill = PatternFill("solid", fgColor="ffff00")
                sheet[column_letter + '9'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            for column in range(imporx, impory):
                column_letter = get_column_letter(column)
                # escribe en la celda
                num_pos = 10
                sheet[column_letter + str(num_pos)] = 'IMPORTE'
                sheet[column_letter + str(num_pos)].font = Font(name='Arial', size=9, bold=True)
                sheet[column_letter + str(10)].fill = PatternFill("solid", fgColor="ffff00")
                sheet[column_letter + str(10)].border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # ---------------------------------- ESTIMADOS --------------------------------------
            acum_est += 1 + aviso
            if contador_est == 1:
                xy = 10
            clave_sub = 0
            for column_estimado in range(imporx-1, imporx):
                column_letter_esti = get_column_letter(column_estimado)
                xd = acum_est + 10
                aviso = 0
                clave_inicial = get_column_letter(2)
                clave_cat = get_column_letter(6)
                acum_clave = 0
                acum_clave_ = 0
                acum_clavexxx = 0 # CATEGORIA
                clave = ''
                for k in b_partida.conceptos_partidas:
                    acum_clave += 1 + clave_sub
                    acum_clavexxx += 1
                    clave_sub = 0
                    clave = sheet[clave_inicial + str(acum_clave + 9)].value
                    categoria = sheet[clave_cat + str(acum_clave + 10)].value

                    # SI LA LINEA ES SUBTOTAL IGNORAR
                    if str(clave) == 'SUBTOTAL':
                        acum_clave_ += 1
                        clave_sub = 1
                        sheet[column_letter_esti + str(acum_clave + 10)].fill = PatternFill("solid", fgColor="bdbdbd")
                        sheet[column_letter_esti + str(acum_clave + 11)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        sheet[column_letter_esti + str(acum_clave + 10)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        # -----------------  SUBTOTAL DE CANTIDAD

                        sheet[column_letter_esti + str(acum_clave + 9)].fill = PatternFill("solid", fgColor="ff7043")
                        sheet[column_letter_esti + str(acum_clave + 9)].border = Border(top=thin, left=thin, right=thin,
                                                                                        bottom=thin)
                        variable4 = acum_clave + 11
                    else:
                        clave_sub = 0

                        if not categoria: # CATEGORIA AGREGAR STILO
                            sheet[column_letter_esti + str(acum_clave + 10)].fill = PatternFill("solid", fgColor="bdbdbd")

                        # SI EL CONCEPTO DEL CICLO ES IGUAL AL DEL CATALOGO, INSERTAR EN ESA POSICION
                        if str(y.clave_linea) == str(clave):
                            sheet[column_letter_esti + str(acum_clave + 9)] = y.estimacion
                            sheet[column_letter_esti + str(acum_clave + 9)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        else:
                            sheet[column_letter_esti + str(acum_clave + 10)].border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # ULTIMO SUBTOTAL DE CANTIDAD
            sheet[column_letter_esti + str(acum_clave + 11)].fill = PatternFill("solid", fgColor="ff7043")
            sheet[column_letter_esti + str(acum_clave + 11)].border = Border(top=thin, left=thin, right=thin,
                                                                            bottom=thin)

            # -------------------------   IMPORTE EJECUTADO ESTIMACION ----------------------------------
            clave_subx = 0

            pos_total2 = 8 + (estimacion_c * 2) + 1 # SUBTOTAL T.IMPORTE
            column_letter_timporte = get_column_letter(pos_total2)
            column_letter_tcantidad = get_column_letter(pos_total2-1)

            for column in range(imporx, impory):
                clave_inicial = get_column_letter(2)
                acum_clavex = 0
                acum_clave_ = 0
                clave = ''
                column_letterc = get_column_letter(column)
                for k in b_partida.conceptos_partidas:
                    acum_clavex += 1 + clave_subx
                    categoria = sheet[clave_cat + str(acum_clavex + 10)].value
                    clave_subx = 0
                    clave = sheet[clave_inicial + str(acum_clavex + 9)].value
                    # SI LA LINEA ES SUBTOTAL IGNORAR
                    if str(clave) == 'SUBTOTAL':
                        acum_clave_ += 1
                        clave_subx = 1
                        # CAT
                        sheet[column_letterc + str(acum_clavex + 10)].fill = PatternFill("solid", fgColor="bdbdbd")
                        sheet[column_letterc + str(acum_clavex + 10)].border = Border(top=thin, left=thin, right=thin,
                                                                                    bottom=thin)
                        # EXCEPCION DE PRIMER CICLO
                        if variable5 > acum_clavex + 8:
                            variable5 = acum_clave_ + 10
                        if int(b_est.idobra) == 1:
                            letra_columna_estimp = get_column_letter(column)
                        else:
                            letra_columna_estimp = get_column_letter(9 + posicion_estimacion)
                        letra_importe = letra_columna_estimp
                        # SUB
                        sheet[column_letterc + str(acum_clavex + 9)] = "=SUM(" + letra_importe + str(variable5) + ":" \
                                                                       + letra_importe + str(acum_clavex + 8) + ")"

                        # SUBTOTAL T.IMPORTE
                        sheet[column_letter_timporte + str(acum_clavex + 9)] = "=SUM(" + column_letter_timporte + str(variable5) + ":" + column_letter_timporte + str(acum_clavex + 8) + ")"

                        sheet[column_letter_timporte + str(acum_clavex + 9)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                        sheet[column_letterc + str(acum_clavex + 9)].fill = PatternFill("solid", fgColor="ff7043")
                        sheet[column_letterc + str(acum_clavex + 9)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                        sheet[column_letterc + str(acum_clavex + 9)].border = Border(top=thin, left=thin, right=thin,
                                                                                        bottom=thin)

                        # SUBTOTAL T.IMPORTE --------------------------------------------------------------------
                        pos_totalsubimport = 8 + (estimacion_c * 2) + 1
                        column_letter_subtimporte = get_column_letter(pos_totalsubimport)
                        sheet[column_letter_subtimporte + str(
                            acum_clavex + 9)] = "=SUM(" + column_letter_subtimporte + str(
                            variable5) + ":" + column_letter_subtimporte + str(acum_clavex + 8) + ")"
                        sheet[column_letter_subtimporte + str(acum_clavex + 9)].fill = PatternFill("solid",
                                                                                                   fgColor="04BDF4")

                        variable5 = acum_clavex + 11
                    else:
                        clave_subx = 0

                        if not categoria: # PRIMERA CATEGORIA STYLE
                            sheet[column_letterc + str(acum_clavex + 10)].fill = PatternFill("solid", fgColor="bdbdbd")
                        # SI EL CONCEPTO DEL CICLO COINCIDE CON EL DEL CATALOGO, INSERTAR EN ESA POSICION
                        if str(y.clave_linea) == str(clave):
                            sheet[column_letterc + str(acum_clavex + 9)] = y.importe_ejecutado
                            sheet[column_letterc + str(acum_clavex + 9)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00' # CURRENCY
                            sheet[column_letterc + str(acum_clavex + 9)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            xy = impory
                        else:
                            sheet[column_letterc + str(acum_clavex + 10)].border = Border(top=thin, left=thin, right=thin, bottom=thin)

                sheet[column_letterc + str(acum_clavex + 13)] = b_est.estimado  # TOTAL ESTIMADO DE CADA ESTIMACION x-x
                sheet[column_letterc + str(acum_clavex + 13)].fill = PatternFill("solid", fgColor="2898B9")
                sheet[column_letterc + str(acum_clavex + 13)].font = Font(color="fafafa", name='Arial', size=11)
                sheet[column_letterc + str(acum_clavex + 13)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'

            # ULTIMO SUBTOTAL DE IMPORTE
            # SUBTOTAL T.IMPORTE
            sheet[column_letter_timporte + str(acum_clavex + 11)] = "=SUM(" + column_letter_timporte + str(variable5) \
                                                                   + ":" + column_letter_timporte + str(acum_clavex + 10) + ")"

            sheet[column_letterc + str(acum_clavex + 11)] = "=SUM(" + letra_importe + str(variable5) + ":" \
                                                           + letra_importe + str(acum_clavex + 10) + ")"
            sheet[column_letterc + str(acum_clavex + 11)].fill = PatternFill("solid", fgColor="ff7043")
            sheet[column_letter_timporte + str(acum_clavex + 11)].fill = PatternFill("solid", fgColor="ff7043")
            sheet[column_letter_timporte + str(acum_clavex + 11)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
            sheet[column_letter_tcantidad + str(acum_clavex + 11)].fill = PatternFill("solid", fgColor="ff7043")
            sheet[column_letter_tcantidad + str(acum_clavex + 11)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet[column_letter_timporte + str(acum_clavex + 11)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet[column_letterc + str(acum_clavex + 11)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
            sheet[column_letterc + str(acum_clavex + 11)].border = Border(top=thin, left=thin, right=thin,
                                                                         bottom=thin)

            pos_concepto3 += 1
            if str(y.precio_unitario) == '0.0':
                y.precio_unitario = ''
            acumy += 1 + aviso
            # ----------------------------------------- T.CANTIDAD y T.IMPORTE------------------------------------------
            pos_total1 = 8 + (estimacion_c * 2)
            pos_total2 = 8 + (estimacion_c * 2) + 1
            for column in range(pos_total1, pos_total2):
                column_letter = get_column_letter(column)
                column_letter_timporte = get_column_letter(pos_total2)
                clave_inicial = get_column_letter(2)
                acum_clavex = 0
                clave_subx = 0
                clave = ''
                acum_linea_total = 0
                for k in b_partida.conceptos_partidas:
                    acum_clavex += 1 + clave_subx
                    categoria = sheet[clave_cat + str(acum_clavex + 10)].value
                    clave = sheet[clave_inicial + str(acum_clavex + 9)].value
                    if str(clave) == 'SUBTOTAL':
                        acum_linea_total += 1
                        clave_subx = 1
                        sheet[column_letter + str(acum_clavex + 9)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        sheet[column_letter + str(acum_clavex + 9)].fill = PatternFill("solid", fgColor="ff7043")
                        sheet[column_letter_timporte + str(acum_clavex + 9)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        sheet[column_letter_timporte + str(acum_clavex + 9)].fill = PatternFill("solid", fgColor="ff7043")
                        sheet[column_letter + str(acum_clavex + 10)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        sheet[column_letter + str(acum_clavex + 10)].fill = PatternFill("solid", fgColor="bdbdbd")
                        sheet[column_letter_timporte + str(acum_clavex + 10)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        sheet[column_letter_timporte + str(acum_clavex + 10)].fill = PatternFill("solid", fgColor="bdbdbd")
                    else:
                        continuacion = ''
                        continuacion_tim = ''
                        letra = 0
                        pasar = 0
                        pos = 8
                        xd2 = acumy + 10
                        column_letterx = ''
                        column_lettery = ''
                        aviso = 0
                        pos_tim = 9
                        xd2_tim = acumy + 10
                        column_letterx_tim = ''
                        column_lettery_tim = ''
                        # CICLO PARA ACUMULAR FORMULAS DE CADA LINEA
                        for i in range(estimacion_c):
                            pasar += 1
                            # pasar los ultimos dos registros
                            if pasar <= (estimacion_c - 2):
                                letra += 2
                                column_lettery2 = get_column_letter((pos + 2) + letra)
                                column_lettery2_tim = get_column_letter((pos_tim + 2) + letra)
                                # STRING ACUMULADO
                                continuacion += ',' + str(column_lettery2 + str(acum_clavex + 9))
                                continuacion_tim += ',' + str(column_lettery2_tim + str(acum_clavex + 9))
                            else:
                                pass
                        column_letterx = get_column_letter(pos)
                        column_lettery = get_column_letter(pos + 2)
                        column_letterx_tim = get_column_letter(pos_tim)
                        column_lettery_tim = get_column_letter(pos_tim + 2)

                        clave_subx = 0

                        if not categoria:  # CATEGORIA AGREGAR STILO
                            sheet[column_letter + str(acum_clavex + 10)].fill = PatternFill("solid", fgColor="bdbdbd")
                            sheet[column_letter + str(acum_clavex + 10)] = ''
                            sheet[column_letter_timporte + str(acum_clavex + 10)].fill = PatternFill("solid",fgColor="bdbdbd")
                            # sheet[column_letter_timporte + str(acum_clavex + 10)] = ''

                        if str(y.clave_linea) == str(clave):
                            # FORMULA PARA SUMAR
                            if estimacion_c == 1:
                                sheet[column_letter + str(acum_clavex + 9)] = y.estimacion
                                sheet[column_letter_timporte + str(acum_clavex + 9)] = y.importe_ejecutado
                                sheet[column_letter + str(acum_clavex + 9)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                                sheet[column_letter_timporte + str(acum_clavex + 9)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                                sheet[column_letter_timporte + str(acum_clavex + 9)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                            else:
                                sheet[column_letter + str(acum_clavex + 9)] = "=SUM(" + column_letterx + str(acum_clavex + 9) \
                                                                              + "," + column_lettery + str(acum_clavex + 9) + continuacion + ")"
                                sheet[column_letter + str(acum_clavex + 9)].border = Border(top=thin, left=thin, right=thin, bottom=thin)

                                sheet[column_letter_timporte + str(acum_clavex + 9)] = "=SUM(" + column_letterx_tim + str(acum_clavex + 9) + "," + \
                                                                                       column_lettery_tim + str(acum_clavex + 9) + continuacion_tim + ")"

                                sheet[column_letter_timporte + str(acum_clavex + 9)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                                sheet[column_letter_timporte + str(acum_clavex + 9)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                        else:
                            sheet[column_letter + str(acum_clavex + 10)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            sheet[column_letter + str(acum_clavex + 9)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            sheet[column_letter_timporte + str(acum_clavex + 9)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            sheet[column_letter_timporte + str(acum_clavex + 10)].border = Border(top=thin, left=thin,
                                                                                                 right=thin,
                                                                                                 bottom=thin)
                xy = impory
                # total_est = acum_clavex + 9
                total_est = 12 + acum_clavex # + acum_linea_total

    # ------------------------------------ TOTALES DE LAS ESTIMACIONES --------------------------------------
    column_letteru = get_column_letter(5)  # TOTAL
    column_letter_tcontrato = get_column_letter(7)  # TOTAL CONTRATO
    column_letter_4 = get_column_letter(pos_total1 + 1)  # TOTAL IMPORTE
    sheet[column_letteru + str(total_est+1)] = 'TOTAL'
    sheet[column_letteru + str(total_est+1)].alignment = Alignment(horizontal="center", vertical="center")
    sheet[column_letter_tcontrato + str(total_est+1)] = total_contrato
    sheet[column_letter_tcontrato + str(total_est+1)].font = Font(color="fafafa", name='Arial', size=13)
    sheet[column_letter_tcontrato + str(total_est+1)].fill = PatternFill("solid", fgColor="2898B9")
    sheet[column_letteru + str(total_est+1)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
    sheet[column_letteru + str(total_est+1)].fill = PatternFill("solid", fgColor="66ADE5")
    sheet[column_letteru + str(total_est+1)].font = Font(color="fafafa", name='Arial', size=13)
    sheet[column_letteru + str(total_est+1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    # sheet[column_letter_3 + str(total_est)].fill = PatternFill("solid", fgColor="66bb6a")
    # sheet[column_letter_3 + str(total_est)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    # sheet[column_letter_4 + str(total_est+1)] = "=SUM(" + column_letter_4 + str(11) + ":" + column_letter_4 + str(total_est-1) + ")"
    sheet[column_letter_4 + str(total_est+1)] = acum_estimado


    sheet[column_letter_4 + str(total_est+1)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
    sheet[column_letter_4 + str(total_est+1)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    sheet[column_letter_4 + str(total_est+1)].fill = PatternFill("solid", fgColor="66bb6a")
    sheet[column_letter_4 + str(total_est+1)].font = Font(color="fafafa", name='Arial', size=11)

    # TOTALES
    for column in range(xy, xy + 1):
        column_letter = get_column_letter(column)
        column_letterq = get_column_letter(column+1)
        # escribe en la celda
        xd2 = 10
        sheet[column_letter + str(xd2)] = 'T.CANTIDAD'
        sheet[column_letter + str(xd2)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet[column_letter + str(xd2)].fill = PatternFill("solid", fgColor="e65100")
        sheet[column_letter + str(xd2)].font = Font(color="fafafa", name='Arial', size=9, bold=True)
        sheet[column_letter + '9'] = 'TOTALES '
        sheet[column_letter + '9'].alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells("" + column_letter + "9:" + column_letterq + "9")
        sheet[column_letterq + '9'].fill = PatternFill("solid", fgColor="e65100")
        sheet[column_letter + '9'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet[column_letter + '9'].fill = PatternFill("solid", fgColor="e65100")
        sheet[column_letter + '9'].font = Font(color="fafafa", name='Arial', size=9, bold=True)
    for column in range(xy + 1, xy + 2):
        column_letter = get_column_letter(column)
        # escribe en la celda
        xdx = 10
        sheet[column_letter + str(xdx)] = 'T.IMPORTE'
        sheet[column_letter + str(xdx)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet[column_letter + str(xdx)].fill = PatternFill("solid", fgColor="e65100")
        sheet[column_letter + str(xdx)].font = Font(color="fafafa", name='Arial', size=9, bold=True)'''

for column in range(7,8):
    column_letter = get_column_letter(column)
    #escribe en la celda
    sheet[column_letter + "7"] = nombre_partida

for column in range(2,3):
    column_letter = get_column_letter(column)
    #escribe en la celda
    sheet[column_letter + "7"] = contratista

# Save the spreadsheet
workbook.save(filename="resultado.xlsx")
